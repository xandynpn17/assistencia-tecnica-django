import csv
from datetime import datetime
from decimal import Decimal, InvalidOperation
import io
import logging
import random
import re
import string

from django.db import transaction
from django.db.models import Sum
from django.utils import timezone

from configuracoes.models import ConfiguracaoSistema, Empresa, FornecedorGarantia, MarcaGarantia

from ..models import (
    ConfiguracaoRateioCustoFixo,
    EstoqueEvento,
    MovimentacaoEstoque,
    PontoOperacional,
    Produto,
    ProdutoHistorico,
    RateioCustoFixoCompetencia,
    ReservaEstoque,
    SaldoEstoquePonto,
    VendaRapidaEstoque,
)
from ..services import (
    ajustar_saldo,
    cancelar_reserva,
    consumir_reservas_ordem,
    converter_reserva,
    devolver_reservas_ordem,
    expirar_reservas_vencidas,
    gerar_codigo_cesto_venda_rapida,
    gerar_codigo_guia_venda_rapida,
    limpar_pre_reservas_antigas,
    normalizar_saldos_produto,
    recalcular_total_produto,
    resumir_cesto_venda_rapida,
    saldo_disponivel,
)

logger = logging.getLogger(__name__)


def _normalizar_saldos_produto(produto):
    normalizar_saldos_produto(produto)


def _recalcular_total_produto(produto):
    recalcular_total_produto(produto)


def _decimal_to_str(valor):
    if valor is None:
        return "0"
    try:
        return str(Decimal(str(valor)))
    except (InvalidOperation, ValueError, TypeError):
        return str(valor)


def _registrar_evento_estoque(evento, *, usuario=None, **dados):
    payload = {
        "evento": evento,
        "usuario_id": getattr(usuario, "id", None) if getattr(usuario, "is_authenticated", False) else None,
        "usuario_nome": getattr(usuario, "username", "") if getattr(usuario, "is_authenticated", False) else "",
    }
    payload.update(dados)
    logger.info("estoque_evento", extra=payload)
    try:
        evento_modelo = {
            "evento": evento,
            "usuario": usuario if getattr(usuario, "is_authenticated", False) else None,
            "quantidade": dados.get("quantidade"),
            "dados": {k: v for k, v in dados.items() if k not in {"quantidade", "produto_id", "ponto_id", "reserva_id", "venda_id", "inventario_id"}},
        }
        if dados.get("produto_id"):
            evento_modelo["produto_id"] = dados.get("produto_id")
        if dados.get("ponto_id"):
            evento_modelo["ponto_operacional_id"] = dados.get("ponto_id")
        if dados.get("reserva_id"):
            evento_modelo["reserva_id"] = dados.get("reserva_id")
        if dados.get("venda_id"):
            evento_modelo["venda_id"] = dados.get("venda_id")
        if dados.get("inventario_id"):
            evento_modelo["inventario_id"] = dados.get("inventario_id")
        EstoqueEvento.objects.create(**evento_modelo)
    except Exception:
        logger.exception("falha_persistir_estoque_evento", extra={"evento": evento})


def _snapshot_produto(produto):
    return {
        "nome": produto.nome or "",
        "sku": produto.sku or "",
        "ean": produto.ean or "",
        "tipo_item": produto.tipo_item or "",
        "preco_final": _decimal_to_str(produto.preco_final),
        "preco_sugerido": _decimal_to_str(produto.preco_sugerido),
        "preco_minimo": _decimal_to_str(produto.preco_minimo),
        "custo_unitario": _decimal_to_str(produto.custo_unitario),
        "custo_operacional": _decimal_to_str(produto.custo_operacional),
        "custo_frete": _decimal_to_str(produto.custo_frete),
        "custo_impostos": _decimal_to_str(produto.custo_impostos),
        "custo_comissao": _decimal_to_str(produto.custo_comissao),
        "custo_marketplace": _decimal_to_str(produto.custo_marketplace),
        "custo_cac": _decimal_to_str(produto.custo_cac),
        "custo_rateio_fixo": _decimal_to_str(getattr(produto, "custo_rateio_fixo", 0)),
        "margem_lucro": _decimal_to_str(produto.margem_lucro),
        "margem_minima": _decimal_to_str(produto.margem_minima),
        "quantidade": int(produto.quantidade or 0),
        "estoque_minimo": int(produto.estoque_minimo or 0),
        "previsao_venda_mensal": int(getattr(produto, "previsao_venda_mensal", 0) or 0),
        "incluir_rateio_custo_fixo": bool(getattr(produto, "incluir_rateio_custo_fixo", False)),
        "permite_comissao_peca": bool(produto.permite_comissao_peca),
        "percentual_comissao_peca": _decimal_to_str(produto.percentual_comissao_peca),
        "bonus_venda": _decimal_to_str(produto.bonus_venda),
    }


def _registrar_historico_produto(produto, *, usuario=None, acao="EDICAO", dados_antes=None, observacao=""):
    ProdutoHistorico.objects.create(
        produto=produto,
        acao=acao,
        usuario=usuario if getattr(usuario, "is_authenticated", False) else None,
        dados_antes=dados_antes or {},
        dados_depois=_snapshot_produto(produto),
        observacao=(observacao or "")[:200],
    )


def _aplicar_estoque_inicial(produto, *, estoque_inicial=0, custo_entrada=None, usuario=None, observacao=""):
    try:
        quantidade = int(estoque_inicial or 0)
    except (TypeError, ValueError):
        quantidade = 0
    if quantidade <= 0 or produto.eh_servico or not produto.ponto_operacional:
        return 0

    custo = Decimal(str(custo_entrada or 0))
    with transaction.atomic():
        ajustar_saldo(produto, produto.ponto_operacional, quantidade)
        MovimentacaoEstoque.objects.create(
            produto=produto,
            tipo="entrada",
            quantidade=quantidade,
            destino=produto.ponto_operacional,
            valor_unitario_custo=custo if custo > 0 else None,
            observacao=(observacao or "Entrada inicial no cadastro do produto.")[:200],
            usuario=usuario if getattr(usuario, "is_authenticated", False) else None,
        )
    return quantidade


def _codigo_reserva():
    while True:
        codigo = "RES-" + "".join(random.choices(string.ascii_uppercase + string.digits, k=8))
        if not ReservaEstoque.objects.filter(codigo_reserva=codigo).exists():
            return codigo


def _codigo_cesto():
    return gerar_codigo_cesto_venda_rapida()


def _codigo_guia():
    return gerar_codigo_guia_venda_rapida()


def _config_sistema():
    return ConfiguracaoSistema.get_configuracao()


def _normalizar_texto(valor):
    texto = (valor or "").strip().lower()
    texto = re.sub(r"\s+", " ", texto)
    return texto


def _resumo_cesto(cesto_codigo):
    return resumir_cesto_venda_rapida(cesto_codigo)


def _initial_produto_from_origem(origem):
    if not origem:
        return {}
    return {
        "nome": origem.nome,
        "tipo_item": origem.tipo_item,
        "modo_preco": origem.modo_preco,
        "descricao": origem.descricao,
        "observacao_interna": origem.observacao_interna,
        "modelos_compativeis": origem.modelos_compativeis,
        "categoria_config": origem.categoria_config_id,
        "categoria": origem.categoria,
        "marca": origem.marca_id,
        "fornecedor_config": origem.fornecedor_config_id,
        "fornecedor_manual": origem.fornecedor_manual,
        "localizacao": origem.localizacao,
        "garantia_peca_dias": origem.garantia_peca_dias,
        "permite_os": origem.permite_os,
        "permite_comissao_peca": origem.permite_comissao_peca,
        "percentual_comissao_peca": origem.percentual_comissao_peca,
        "bonus_venda": origem.bonus_venda,
        "servicos_compativeis": list(origem.servicos_compativeis.values_list("id", flat=True)),
        "custo_unitario": origem.custo_unitario,
        "custo_operacional": origem.custo_operacional,
        "custo_frete": origem.custo_frete,
        "custo_impostos": origem.custo_impostos,
        "custo_comissao": origem.custo_comissao,
        "custo_marketplace": origem.custo_marketplace,
        "custo_cac": origem.custo_cac,
        "previsao_venda_mensal": getattr(origem, "previsao_venda_mensal", 0),
        "incluir_rateio_custo_fixo": getattr(origem, "incluir_rateio_custo_fixo", False),
        "custo_medio": origem.custo_medio,
        "margem_lucro": origem.margem_lucro,
        "margem_minima": origem.margem_minima,
        "taxa_cartao": origem.taxa_cartao,
        "usar_aliquota_manual": origem.usar_aliquota_manual,
        "aliquota_manual": origem.aliquota_manual,
        "icms": origem.icms,
        "ipi": origem.ipi,
        "pis": origem.pis,
        "cofins": origem.cofins,
        "pis_cofins": origem.pis_cofins,
        "preco_final": origem.preco_final,
        "estoque_minimo": origem.estoque_minimo,
        "ativo": origem.ativo,
        "data_entrada": timezone.localdate(),
        "ponto_operacional": origem.ponto_operacional_id,
    }


def _ler_arquivo_importacao_produtos(uploaded_file):
    nome = (getattr(uploaded_file, "name", "") or "").lower()
    if nome.endswith(".csv"):
        data = uploaded_file.read()
        texto = data.decode("utf-8-sig", errors="ignore")
        try:
            dialect = csv.Sniffer().sniff(texto[:4096], delimiters=",;")
        except csv.Error:
            dialect = csv.excel
        leitor = csv.DictReader(io.StringIO(texto), dialect=dialect)
        rows = []
        for row in leitor:
            normalizada = {(k or "").strip(): (v or "").strip() for k, v in row.items()}
            if any(normalizada.values()):
                rows.append(normalizada)
        return rows

    if nome.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Importacao XLSX requer openpyxl instalado.") from exc

        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active
        headers = []
        rows = []
        for idx, line in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx == 1:
                headers = [str(c or "").strip() for c in line]
                continue
            registro = {}
            for pos, header in enumerate(headers):
                if not header:
                    continue
                registro[header] = str(line[pos] if pos < len(line) and line[pos] is not None else "").strip()
            if any(registro.values()):
                rows.append(registro)
        return rows
    raise ValueError("Formato invalido. Envie arquivo CSV ou XLSX.")


def _normalizar_linha_importacao(linha):
    def _val(key, default=""):
        return (linha.get(key) or linha.get(key.lower()) or "").strip() or default

    return {
        "nome": _val("nome"),
        "sku": _val("sku"),
        "ean": _val("ean"),
        "tipo_item": (_val("tipo_item", "produto") or "produto").lower(),
        "categoria": _val("categoria"),
        "marca_nome": _val("marca"),
        "fornecedor_nome": _val("fornecedor"),
        "preco_final": _val("preco_final", "0"),
        "custo_unitario": _val("custo_unitario", "0"),
        "estoque_minimo": _val("estoque_minimo", "0"),
        "estoque_inicial": _val("estoque_inicial", "0"),
    }


def _contexto_rateio_produto(produto=None, empresa=None):
    competencia = timezone.localdate().replace(day=1)
    configuracao = ConfiguracaoRateioCustoFixo.get_solo()
    total_fixos = Decimal("0.00")
    try:
        from caixa.models import CustoFixoMensal

        total_fixos = (
            CustoFixoMensal.objects.filter(competencia=competencia, ativo=True)
            .exclude(status="cancelado")
            .aggregate(total=Sum("valor_previsto"))["total"]
            or Decimal("0.00")
        )
    except (ImportError, ModuleNotFoundError):
        total_fixos = Decimal("0.00")

    produtos_base = Produto.objects.ativos().nao_servicos().filter(incluir_rateio_custo_fixo=True, previsao_venda_mensal__gt=0)
    if empresa:
        produtos_base = produtos_base.filter(empresa=empresa)
    if produto and getattr(produto, "pk", None):
        produtos_base = produtos_base.exclude(pk=produto.pk)

    total_base_rateio = Decimal("0.00")
    for produto_base in produtos_base:
        total_base_rateio += produto_base.base_rateio_custo_fixo(criterio=configuracao.criterio_rateio)

    return {
        "competencia": competencia,
        "criterio_rateio": configuracao.criterio_rateio,
        "criterio_rateio_display": configuracao.get_criterio_rateio_display(),
        "total_fixos": total_fixos,
        "total_base_rateio": total_base_rateio,
    }


def _resumo_rateio_atual(empresa=None):
    competencia = timezone.localdate().replace(day=1)
    configuracao = ConfiguracaoRateioCustoFixo.get_solo()
    total_fixos = Decimal("0.00")
    try:
        from caixa.models import CustoFixoMensal

        total_fixos = (
            CustoFixoMensal.objects.filter(competencia=competencia, ativo=True)
            .exclude(status="cancelado")
            .aggregate(total=Sum("valor_previsto"))["total"]
            or Decimal("0.00")
        )
    except (ImportError, ModuleNotFoundError):
        total_fixos = Decimal("0.00")

    produtos = list(
        Produto.objects.ativos().nao_servicos().filter(incluir_rateio_custo_fixo=True, previsao_venda_mensal__gt=0).order_by("nome")
    )
    if empresa:
        produtos = [produto for produto in produtos if produto.empresa_id == empresa.id]
    detalhes = []
    total_base = Decimal("0.00")
    realizado_por_produto = RateioCustoFixoCompetencia._realizado_por_produto(competencia)
    for produto in produtos:
        base_rateio = produto.base_rateio_custo_fixo(criterio=configuracao.criterio_rateio)
        if base_rateio <= 0:
            continue
        custo_unitario = produto.calcular_rateio_custo_fixo_unitario(competencia=competencia, criterio_override=configuracao.criterio_rateio)
        custo_total = Decimal(str(produto.previsao_venda_mensal or 0)) * custo_unitario
        realizado = realizado_por_produto.get(produto.id, {})
        total_base += base_rateio
        detalhes.append(
            {
                "produto": produto,
                "base_rateio": base_rateio,
                "custo_rateio_unitario": custo_unitario,
                "custo_rateio_total": custo_total,
                "quantidade_realizada": int(realizado.get("quantidade", 0) or 0),
                "faturamento_realizado": Decimal(str(realizado.get("faturamento", 0) or 0)),
                "margem_realizada": Decimal(str(realizado.get("margem", 0) or 0)),
            }
        )

    for detalhe in detalhes:
        detalhe["participacao_percentual"] = Decimal("0.00") if total_base <= 0 else (detalhe["base_rateio"] / total_base) * Decimal("100")
        previsao = int(detalhe["produto"].previsao_venda_mensal or 0)
        detalhe["percentual_realizado"] = Decimal("0.00")
        if previsao > 0:
            detalhe["percentual_realizado"] = (Decimal(str(detalhe["quantidade_realizada"])) / Decimal(str(previsao))) * Decimal("100")

    detalhes.sort(key=lambda item: (item["custo_rateio_total"], item["produto"].nome), reverse=True)
    snapshots = list(RateioCustoFixoCompetencia.objects.select_related("gerado_por").prefetch_related("itens").order_by("-competencia")[:6])
    historico_competencias = []
    for snapshot in snapshots:
        previstos = sum((int(item.previsao_venda_mensal or 0) for item in snapshot.itens.all()), 0)
        realizados = sum((int(item.quantidade_realizada or 0) for item in snapshot.itens.all()), 0)
        faturamento = sum((item.faturamento_realizado for item in snapshot.itens.all()), Decimal("0.00"))
        margem = sum((item.margem_realizada for item in snapshot.itens.all()), Decimal("0.00"))
        percentual_realizado = Decimal("0.00")
        if previstos > 0:
            percentual_realizado = (Decimal(str(realizados)) / Decimal(str(previstos))) * Decimal("100")
        historico_competencias.append(
            {
                "snapshot": snapshot,
                "previstos": previstos,
                "realizados": realizados,
                "faturamento": faturamento,
                "margem": margem,
                "percentual_realizado": percentual_realizado,
            }
        )

    total_previsto = sum((int(item["produto"].previsao_venda_mensal or 0) for item in detalhes), 0)
    total_realizado = sum((int(item["quantidade_realizada"] or 0) for item in detalhes), 0)
    percentual_realizado_total = Decimal("0.00")
    if total_previsto > 0:
        percentual_realizado_total = (Decimal(str(total_realizado)) / Decimal(str(total_previsto))) * Decimal("100")
    return {
        "competencia": competencia,
        "configuracao": configuracao,
        "total_fixos": total_fixos,
        "total_base_rateio": total_base,
        "total_produtos": len(detalhes),
        "total_previsto": total_previsto,
        "total_realizado": total_realizado,
        "percentual_realizado_total": percentual_realizado_total,
        "total_faturamento_realizado": sum((item["faturamento_realizado"] for item in detalhes), Decimal("0.00")),
        "total_margem_realizada": sum((item["margem_realizada"] for item in detalhes), Decimal("0.00")),
        "produtos": detalhes[:20],
        "snapshots": snapshots,
        "historico_competencias": historico_competencias,
        "snapshot_atual": RateioCustoFixoCompetencia.objects.filter(competencia=competencia).first(),
    }


__all__ = [
    "Empresa",
    "FornecedorGarantia",
    "MarcaGarantia",
    "_aplicar_estoque_inicial",
    "_codigo_cesto",
    "_codigo_guia",
    "_codigo_reserva",
    "_config_sistema",
    "_contexto_rateio_produto",
    "_initial_produto_from_origem",
    "_ler_arquivo_importacao_produtos",
    "_normalizar_linha_importacao",
    "_normalizar_saldos_produto",
    "_normalizar_texto",
    "_recalcular_total_produto",
    "_registrar_evento_estoque",
    "_registrar_historico_produto",
    "_resumo_cesto",
    "_resumo_rateio_atual",
    "_snapshot_produto",
    "ajustar_saldo",
    "cancelar_reserva",
    "consumir_reservas_ordem",
    "converter_reserva",
    "datetime",
    "devolver_reservas_ordem",
    "expirar_reservas_vencidas",
    "limpar_pre_reservas_antigas",
    "logger",
    "saldo_disponivel",
]

