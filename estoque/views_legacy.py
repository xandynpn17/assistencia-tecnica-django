import csv
from datetime import datetime, timedelta
from decimal import Decimal
import io
import logging
import random
import re
import string

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from configuracoes.permissions import (
    CAIXA_OPERATIONAL_ROLES,
    ORDER_ROLES,
    STOCK_MANAGE_ROLES,
    STOCK_VIEW_ROLES,
    has_role,
    role_required,
)
from configuracoes.models import ConfiguracaoSistema, Empresa, FornecedorGarantia, MarcaGarantia

from .forms import (
    ConfiguracaoRateioCustoFixoForm,
    GerarSnapshotRateioForm,
    MovimentacaoEstoqueForm,
    PontoOperacionalForm,
    ProdutoEquivalenteForm,
    ProdutoForm,
    ProdutoKitItemForm,
    ProdutoPrecoTabelaForm,
    TabelaPrecoForm,
    UbicacaoEstoqueForm,
)
from .models import (
    ConfiguracaoRateioCustoFixo,
    InventarioEstoque,
    ItemInventarioEstoque,
    MovimentacaoEstoque,
    PontoOperacional,
    Produto,
    ProdutoHistorico,
    ProdutoEquivalente,
    ProdutoKitItem,
    ProdutoPrecoTabela,
    RateioCustoFixoCompetencia,
    ReservaEstoque,
    SaldoEstoquePonto,
    ServicoReferencia,
    TabelaPreco,
    UbicacaoEstoque,
    VendaRapidaEstoque,
)
from .services import (
    ajustar_saldo,
    cancelar_reserva,
    consumir_reservas_ordem,
    converter_reserva,
    devolver_reservas_ordem,
    expirar_reservas_vencidas,
    limpar_pre_reservas_antigas,
    recalcular_total_produto,
    saldo_disponivel,
)

logger = logging.getLogger(__name__)


def _normalizar_saldos_produto(produto):
    if not produto.saldos_por_ponto.exists() and produto.ponto_operacional and produto.quantidade:
        SaldoEstoquePonto.objects.create(
            produto=produto,
            ponto_operacional=produto.ponto_operacional,
            quantidade=produto.quantidade,
        )


def _recalcular_total_produto(produto):
    recalcular_total_produto(produto)


def _decimal_to_str(valor):
    if valor is None:
        return "0"
    try:
        return str(Decimal(str(valor)))
    except Exception:
        return str(valor)


def _snapshot_produto(produto):
    return {
        "nome": produto.nome or "",
        "sku": produto.sku or "",
        "ean": produto.ean or "",
        "tipo_item": produto.tipo_item or "",
        "preco_final": _decimal_to_str(produto.preco_final),
        "preco_sugerido": _decimal_to_str(produto.preco_sugerido),
        "preco_minimo": _decimal_to_str(produto.preco_minimo),
        "custo_unitario": _decimal_to_str(produto.custo_unitario),
        "custo_operacional": _decimal_to_str(produto.custo_operacional),
        "custo_frete": _decimal_to_str(produto.custo_frete),
        "custo_impostos": _decimal_to_str(produto.custo_impostos),
        "custo_comissao": _decimal_to_str(produto.custo_comissao),
        "custo_marketplace": _decimal_to_str(produto.custo_marketplace),
        "custo_cac": _decimal_to_str(produto.custo_cac),
        "custo_rateio_fixo": _decimal_to_str(getattr(produto, "custo_rateio_fixo", 0)),
        "margem_lucro": _decimal_to_str(produto.margem_lucro),
        "margem_minima": _decimal_to_str(produto.margem_minima),
        "quantidade": int(produto.quantidade or 0),
        "estoque_minimo": int(produto.estoque_minimo or 0),
        "previsao_venda_mensal": int(getattr(produto, "previsao_venda_mensal", 0) or 0),
        "incluir_rateio_custo_fixo": bool(getattr(produto, "incluir_rateio_custo_fixo", False)),
        "permite_comissao_peca": bool(produto.permite_comissao_peca),
        "percentual_comissao_peca": _decimal_to_str(produto.percentual_comissao_peca),
        "bonus_venda": _decimal_to_str(produto.bonus_venda),
    }


def _registrar_historico_produto(produto, *, usuario=None, acao="EDICAO", dados_antes=None, observacao=""):
    ProdutoHistorico.objects.create(
        produto=produto,
        acao=acao,
        usuario=usuario if getattr(usuario, "is_authenticated", False) else None,
        dados_antes=dados_antes or {},
        dados_depois=_snapshot_produto(produto),
        observacao=(observacao or "")[:200],
    )


def _aplicar_estoque_inicial(produto, *, estoque_inicial=0, custo_entrada=None, usuario=None, observacao=""):
    try:
        quantidade = int(estoque_inicial or 0)
    except (TypeError, ValueError):
        quantidade = 0
    if quantidade <= 0 or produto.is_servico or not produto.ponto_operacional:
        return 0

    custo = Decimal(str(custo_entrada or 0))
    with transaction.atomic():
        ajustar_saldo(produto, produto.ponto_operacional, quantidade)
        MovimentacaoEstoque.objects.create(
            produto=produto,
            tipo="entrada",
            quantidade=quantidade,
            destino=produto.ponto_operacional,
            valor_unitario_custo=custo if custo > 0 else None,
            observacao=(observacao or "Entrada inicial no cadastro do produto.")[:200],
            usuario=usuario if getattr(usuario, "is_authenticated", False) else None,
        )
    return quantidade


def _codigo_reserva():
    while True:
        codigo = "RES-" + "".join(random.choices(string.ascii_uppercase + string.digits, k=8))
        if not ReservaEstoque.objects.filter(codigo_reserva=codigo).exists():
            return codigo


def _codigo_cesto():
    while True:
        codigo = "CES-" + "".join(random.choices(string.digits, k=8))
        if not VendaRapidaEstoque.objects.filter(cesto_codigo=codigo).exists():
            return codigo


def _codigo_guia():
    while True:
        codigo = "GUIA-" + "".join(random.choices(string.digits, k=8))
        if not VendaRapidaEstoque.objects.filter(guia_pagamento=codigo).exists():
            return codigo


def _config_sistema():
    return ConfiguracaoSistema.get_configuracao()


def _normalizar_texto(valor):
    texto = (valor or "").strip().lower()
    texto = re.sub(r"\s+", " ", texto)
    return texto


def _resumo_cesto(cesto_codigo):
    vendas = list(
        VendaRapidaEstoque.objects.select_related("produto", "ponto_operacional")
        .filter(cesto_codigo=cesto_codigo, status="pre_reserva")
        .order_by("-id")
    )
    total = sum((v.valor_total for v in vendas), Decimal("0.00"))
    guia = ""
    for v in vendas:
        if v.guia_pagamento:
            guia = v.guia_pagamento
            break
    return {
        "ok": True,
        "cesto_codigo": cesto_codigo,
        "guia": guia,
        "itens": [
            {
                "id": v.id,
                "produto": v.produto.nome,
                "ponto": v.ponto_operacional.codigo,
                "quantidade": v.quantidade,
                "valor_unitario": float(v.valor_unitario),
                "valor_total": float(v.valor_total),
                "vendedor": v.funcionario_numero,
            }
            for v in vendas
        ],
        "total": float(total),
    }


def _initial_produto_from_origem(origem):
    if not origem:
        return {}
    return {
        "nome": origem.nome,
        "tipo_item": origem.tipo_item,
        "modo_preco": origem.modo_preco,
        "descricao": origem.descricao,
        "observacao_interna": origem.observacao_interna,
        "modelos_compativeis": origem.modelos_compativeis,
        "categoria_config": origem.categoria_config_id,
        "categoria": origem.categoria,
        "marca": origem.marca_id,
        "fornecedor_config": origem.fornecedor_config_id,
        "fornecedor_manual": origem.fornecedor_manual,
        "localizacao": origem.localizacao,
        "garantia_peca_dias": origem.garantia_peca_dias,
        "permite_os": origem.permite_os,
        "permite_comissao_peca": origem.permite_comissao_peca,
        "percentual_comissao_peca": origem.percentual_comissao_peca,
        "bonus_venda": origem.bonus_venda,
        "servicos_compativeis": list(origem.servicos_compativeis.values_list("id", flat=True)),
        "custo_unitario": origem.custo_unitario,
        "custo_operacional": origem.custo_operacional,
        "custo_frete": origem.custo_frete,
        "custo_impostos": origem.custo_impostos,
        "custo_comissao": origem.custo_comissao,
        "custo_marketplace": origem.custo_marketplace,
        "custo_cac": origem.custo_cac,
        "previsao_venda_mensal": getattr(origem, "previsao_venda_mensal", 0),
        "incluir_rateio_custo_fixo": getattr(origem, "incluir_rateio_custo_fixo", False),
        "custo_medio": origem.custo_medio,
        "margem_lucro": origem.margem_lucro,
        "margem_minima": origem.margem_minima,
        "taxa_cartao": origem.taxa_cartao,
        "usar_aliquota_manual": origem.usar_aliquota_manual,
        "aliquota_manual": origem.aliquota_manual,
        "icms": origem.icms,
        "ipi": origem.ipi,
        "pis": origem.pis,
        "cofins": origem.cofins,
        "pis_cofins": origem.pis_cofins,
        "preco_final": origem.preco_final,
        "estoque_minimo": origem.estoque_minimo,
        "ativo": origem.ativo,
        "data_entrada": timezone.localdate(),
        "ponto_operacional": origem.ponto_operacional_id,
    }


def _contexto_rateio_produto(produto=None):
    competencia = timezone.localdate().replace(day=1)
    configuracao = ConfiguracaoRateioCustoFixo.get_solo()
    total_fixos = Decimal("0.00")
    try:
        from caixa.models import CustoFixoMensal

        total_fixos = (
            CustoFixoMensal.objects.filter(competencia=competencia, ativo=True)
            .exclude(status="cancelado")
            .aggregate(total=Sum("valor_previsto"))["total"]
            or Decimal("0.00")
        )
    except Exception:
        total_fixos = Decimal("0.00")

    produtos_base = Produto.objects.filter(
        ativo=True,
        is_servico=False,
        incluir_rateio_custo_fixo=True,
        previsao_venda_mensal__gt=0,
    )
    if produto and getattr(produto, "pk", None):
        produtos_base = produtos_base.exclude(pk=produto.pk)

    total_base_rateio = Decimal("0.00")
    for produto_base in produtos_base:
        total_base_rateio += produto_base.base_rateio_custo_fixo(criterio=configuracao.criterio_rateio)

    return {
        "competencia": competencia,
        "criterio_rateio": configuracao.criterio_rateio,
        "criterio_rateio_display": configuracao.get_criterio_rateio_display(),
        "total_fixos": total_fixos,
        "total_base_rateio": total_base_rateio,
    }


def _resumo_rateio_atual():
    competencia = timezone.localdate().replace(day=1)
    configuracao = ConfiguracaoRateioCustoFixo.get_solo()
    total_fixos = Decimal("0.00")
    try:
        from caixa.models import CustoFixoMensal

        total_fixos = (
            CustoFixoMensal.objects.filter(competencia=competencia, ativo=True)
            .exclude(status="cancelado")
            .aggregate(total=Sum("valor_previsto"))["total"]
            or Decimal("0.00")
        )
    except Exception:
        total_fixos = Decimal("0.00")

    produtos = list(
        Produto.objects.filter(
            ativo=True,
            is_servico=False,
            incluir_rateio_custo_fixo=True,
            previsao_venda_mensal__gt=0,
        ).order_by("nome")
    )

    detalhes = []
    total_base = Decimal("0.00")
    realizado_por_produto = RateioCustoFixoCompetencia._realizado_por_produto(competencia)
    for produto in produtos:
        base_rateio = produto.base_rateio_custo_fixo(criterio=configuracao.criterio_rateio)
        if base_rateio <= 0:
            continue
        custo_unitario = produto.calcular_rateio_custo_fixo_unitario(
            competencia=competencia,
            criterio_override=configuracao.criterio_rateio,
        )
        custo_total = Decimal(str(produto.previsao_venda_mensal or 0)) * custo_unitario
        realizado = realizado_por_produto.get(produto.id, {})
        total_base += base_rateio
        detalhes.append(
            {
                "produto": produto,
                "base_rateio": base_rateio,
                "custo_rateio_unitario": custo_unitario,
                "custo_rateio_total": custo_total,
                "quantidade_realizada": int(realizado.get("quantidade", 0) or 0),
                "faturamento_realizado": Decimal(str(realizado.get("faturamento", 0) or 0)),
                "margem_realizada": Decimal(str(realizado.get("margem", 0) or 0)),
            }
        )

    for detalhe in detalhes:
        detalhe["participacao_percentual"] = (
            Decimal("0.00") if total_base <= 0 else (detalhe["base_rateio"] / total_base) * Decimal("100")
        )
        previsao = int(detalhe["produto"].previsao_venda_mensal or 0)
        detalhe["percentual_realizado"] = Decimal("0.00")
        if previsao > 0:
            detalhe["percentual_realizado"] = (Decimal(str(detalhe["quantidade_realizada"])) / Decimal(str(previsao))) * Decimal("100")

    detalhes.sort(key=lambda item: (item["custo_rateio_total"], item["produto"].nome), reverse=True)
    snapshots = list(
        RateioCustoFixoCompetencia.objects.select_related("gerado_por")
        .prefetch_related("itens")
        .order_by("-competencia")[:6]
    )
    historico_competencias = []
    for snapshot in snapshots:
        previstos = sum((int(item.previsao_venda_mensal or 0) for item in snapshot.itens.all()), 0)
        realizados = sum((int(item.quantidade_realizada or 0) for item in snapshot.itens.all()), 0)
        faturamento = sum((item.faturamento_realizado for item in snapshot.itens.all()), Decimal("0.00"))
        margem = sum((item.margem_realizada for item in snapshot.itens.all()), Decimal("0.00"))
        percentual_realizado = Decimal("0.00")
        if previstos > 0:
            percentual_realizado = (Decimal(str(realizados)) / Decimal(str(previstos))) * Decimal("100")
        historico_competencias.append(
            {
                "snapshot": snapshot,
                "previstos": previstos,
                "realizados": realizados,
                "faturamento": faturamento,
                "margem": margem,
                "percentual_realizado": percentual_realizado,
            }
        )

    total_previsto = sum((int(item["produto"].previsao_venda_mensal or 0) for item in detalhes), 0)
    total_realizado = sum((int(item["quantidade_realizada"] or 0) for item in detalhes), 0)
    percentual_realizado_total = Decimal("0.00")
    if total_previsto > 0:
        percentual_realizado_total = (Decimal(str(total_realizado)) / Decimal(str(total_previsto))) * Decimal("100")
    return {
        "competencia": competencia,
        "configuracao": configuracao,
        "total_fixos": total_fixos,
        "total_base_rateio": total_base,
        "total_produtos": len(detalhes),
        "total_previsto": total_previsto,
        "total_realizado": total_realizado,
        "percentual_realizado_total": percentual_realizado_total,
        "total_faturamento_realizado": sum((item["faturamento_realizado"] for item in detalhes), Decimal("0.00")),
        "total_margem_realizada": sum((item["margem_realizada"] for item in detalhes), Decimal("0.00")),
        "produtos": detalhes[:20],
        "snapshots": snapshots,
        "historico_competencias": historico_competencias,
        "snapshot_atual": RateioCustoFixoCompetencia.objects.filter(competencia=competencia).first(),
    }


@role_required(STOCK_VIEW_ROLES)
def buscar_produtos(request):
    messages.info(request, "A tela 'Buscar Produto' foi descontinuada. Use 'Consulta de Artigos'.")
    return redirect("estoque:consulta_artigos")


@role_required(STOCK_VIEW_ROLES)
def lista_produtos(request):
    filtro = request.GET.get("tipo", "todos")
    ponto_id = request.GET.get("ponto")
    page_number = request.GET.get("page")

    if filtro == "servicos":
        produtos = Produto.objects.filter(ativo=True, is_servico=True)
    elif filtro == "produtos":
        produtos = Produto.objects.filter(ativo=True, is_servico=False)
    else:
        produtos = Produto.objects.filter(ativo=True)

    if ponto_id:
        produtos = produtos.filter(ponto_operacional_id=ponto_id)

    produtos = produtos.select_related("ponto_operacional", "categoria_config", "marca", "fornecedor_config").order_by("nome")
    produtos_page = Paginator(produtos, 30).get_page(page_number)

    context = {
        "produtos": produtos_page,
        "produtos_page": produtos_page,
        "pontos": PontoOperacional.objects.filter(ativo=True),
        "menu_app": "estoque",
        "menu_sub": "lista_produtos",
        "filtro": filtro,
        "ponto_filtro": ponto_id or "",
    }
    return render(request, "estoque/lista_produtos.html", context)


@role_required(STOCK_MANAGE_ROLES)
def criar_produto(request):
    ultimo = Produto.objects.order_by("-id").first()
    initial = {}
    if ultimo:
        initial = {
            "icms": ultimo.icms,
            "ipi": ultimo.ipi,
            "pis_cofins": ultimo.pis_cofins,
            "margem_lucro": ultimo.margem_lucro,
            "custo_operacional": ultimo.custo_operacional,
            "custo_cac": getattr(ultimo, "custo_cac", 0),
            "previsao_venda_mensal": getattr(ultimo, "previsao_venda_mensal", 0),
            "incluir_rateio_custo_fixo": getattr(ultimo, "incluir_rateio_custo_fixo", False),
            "ponto_operacional": ultimo.ponto_operacional_id,
        }
    duplicar_id = request.GET.get("duplicar")
    produto_origem = None
    if duplicar_id and str(duplicar_id).isdigit():
        produto_origem = Produto.objects.filter(id=int(duplicar_id), ativo=True).first()
        if produto_origem:
            initial.update(_initial_produto_from_origem(produto_origem))
            initial["nome"] = f"{produto_origem.nome} (cópia)"
            initial["sku"] = ""
            initial["ean"] = ""
            initial["estoque_inicial"] = 0

    if request.method == "POST":
        form = ProdutoForm(request.POST, request.FILES)
        if form.is_valid():
            produto = form.save()
            estoque_inicial = form.cleaned_data.get("estoque_inicial") or 0
            custo_entrada = form.cleaned_data.get("custo_entrada_inicial")
            _aplicar_estoque_inicial(
                produto,
                estoque_inicial=estoque_inicial,
                custo_entrada=custo_entrada,
                usuario=request.user,
                observacao="Entrada inicial gerada no cadastro do produto.",
            )
            _normalizar_saldos_produto(produto)
            _recalcular_total_produto(produto)
            _registrar_historico_produto(
                produto,
                usuario=request.user,
                acao="DUPLICACAO" if produto_origem else "CRIACAO",
                dados_antes=_snapshot_produto(produto_origem) if produto_origem else {},
                observacao="Cadastro inicial de produto.",
            )
            if "_save_and_new" in request.POST:
                messages.success(request, "Produto cadastrado. Pronto para incluir o próximo.")
                return redirect("estoque:criar_produto")
            messages.success(request, "Produto cadastrado com sucesso.")
            return redirect("estoque:lista_produtos")
    else:
        form = ProdutoForm(initial=initial)

    context = {
        "form": form,
        "menu_app": "estoque",
        "menu_sub": "criar_produto",
        "produto_origem": produto_origem,
        "modo_edicao": False,
        "rateio_context": _contexto_rateio_produto(produto_origem),
    }
    context["empresa"] = Empresa.objects.first()
    return render(request, "estoque/form_produto.html", context)


@role_required(STOCK_MANAGE_ROLES)
def editar_produto(request, produto_id):
    produto = get_object_or_404(Produto, id=produto_id)
    if request.method == "POST":
        snapshot_antes = _snapshot_produto(produto)
        form = ProdutoForm(request.POST, request.FILES, instance=produto)
        if form.is_valid():
            produto = form.save()
            estoque_inicial = form.cleaned_data.get("estoque_inicial") or 0
            custo_entrada = form.cleaned_data.get("custo_entrada_inicial")
            _aplicar_estoque_inicial(
                produto,
                estoque_inicial=estoque_inicial,
                custo_entrada=custo_entrada,
                usuario=request.user,
                observacao="Entrada manual adicional no cadastro do produto.",
            )
            _normalizar_saldos_produto(produto)
            _recalcular_total_produto(produto)
            _registrar_historico_produto(
                produto,
                usuario=request.user,
                acao="EDICAO",
                dados_antes=snapshot_antes,
                observacao="Atualização de cadastro do produto.",
            )
            messages.success(request, "Produto atualizado com sucesso!")
            if "_save_and_new" in request.POST:
                return redirect("estoque:criar_produto")
            return redirect("estoque:lista_produtos")
    else:
        form = ProdutoForm(instance=produto)

    return render(
        request,
        "estoque/form_produto.html",
        {
            "form": form,
            "produto": produto,
            "empresa": Empresa.objects.first(),
            "menu_app": "estoque",
            "menu_sub": "lista_produtos",
            "modo_edicao": True,
            "rateio_context": _contexto_rateio_produto(produto),
        },
    )


@role_required(STOCK_MANAGE_ROLES)
def duplicar_produto(request, produto_id):
    return redirect(f"{reverse('estoque:criar_produto')}?duplicar={produto_id}")


def _ler_arquivo_importacao_produtos(uploaded_file):
    nome = (getattr(uploaded_file, "name", "") or "").lower()
    if nome.endswith(".csv"):
        data = uploaded_file.read()
        texto = data.decode("utf-8-sig", errors="ignore")
        try:
            dialect = csv.Sniffer().sniff(texto[:4096], delimiters=",;")
        except Exception:
            dialect = csv.excel
        leitor = csv.DictReader(io.StringIO(texto), dialect=dialect)
        rows = []
        for row in leitor:
            normalizada = {(k or "").strip(): (v or "").strip() for k, v in row.items()}
            if any(normalizada.values()):
                rows.append(normalizada)
        return rows

    if nome.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise ValueError("Importação XLSX requer openpyxl instalado.") from exc

        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active
        headers = []
        rows = []
        for idx, line in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx == 1:
                headers = [str(c or "").strip() for c in line]
                continue
            registro = {}
            for pos, header in enumerate(headers):
                if not header:
                    continue
                registro[header] = str(line[pos] if pos < len(line) and line[pos] is not None else "").strip()
            if any(registro.values()):
                rows.append(registro)
        return rows
    raise ValueError("Formato inválido. Envie arquivo CSV ou XLSX.")


def _normalizar_linha_importacao(linha):
    def _val(key, default=""):
        return (linha.get(key) or linha.get(key.lower()) or "").strip() or default

    return {
        "nome": _val("nome"),
        "sku": _val("sku"),
        "ean": _val("ean"),
        "tipo_item": (_val("tipo_item", "produto") or "produto").lower(),
        "categoria": _val("categoria"),
        "marca_nome": _val("marca"),
        "fornecedor_nome": _val("fornecedor"),
        "preco_final": _val("preco_final", "0"),
        "custo_unitario": _val("custo_unitario", "0"),
        "estoque_minimo": _val("estoque_minimo", "0"),
        "estoque_inicial": _val("estoque_inicial", "0"),
    }


@role_required(STOCK_MANAGE_ROLES)
def importar_produtos(request):
    preview = []
    erros = []
    importados = 0
    if request.method == "POST":
        arquivo = request.FILES.get("arquivo")
        acao = (request.POST.get("acao") or "validar").strip().lower()
        if not arquivo:
            messages.error(request, "Selecione um arquivo CSV ou XLSX para importar.")
            return redirect("estoque:importar_produtos")
        try:
            linhas = _ler_arquivo_importacao_produtos(arquivo)
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect("estoque:importar_produtos")

        normalizadas = []
        nomes_arquivo = set()
        eans_arquivo = set()
        for idx, linha in enumerate(linhas, start=2):
            row = _normalizar_linha_importacao(linha)
            row["linha"] = idx
            row["erros"] = []
            if not row["nome"]:
                row["erros"].append("Nome obrigatório.")
            nome_key = _normalizar_texto(row["nome"])
            if nome_key in nomes_arquivo and nome_key:
                row["erros"].append("Nome duplicado no arquivo.")
            elif nome_key:
                nomes_arquivo.add(nome_key)
            if row["tipo_item"] not in {"produto", "peca", "consumivel", "servico"}:
                row["erros"].append("Tipo de item inválido.")
            ean_limpo = "".join(ch for ch in (row["ean"] or "") if ch.isdigit())
            row["ean"] = ean_limpo
            if ean_limpo and len(ean_limpo) != 13:
                row["erros"].append("EAN deve ter 13 dígitos.")
            if ean_limpo:
                if ean_limpo in eans_arquivo:
                    row["erros"].append("EAN duplicado no arquivo.")
                else:
                    eans_arquivo.add(ean_limpo)
            try:
                row["preco_final_dec"] = Decimal(str(row["preco_final"] or "0"))
                if row["preco_final_dec"] < 0:
                    row["erros"].append("Preço final não pode ser negativo.")
            except Exception:
                row["erros"].append("Preço final inválido.")
                row["preco_final_dec"] = Decimal("0")
            try:
                row["custo_unitario_dec"] = Decimal(str(row["custo_unitario"] or "0"))
                if row["custo_unitario_dec"] < 0:
                    row["erros"].append("Custo unitário não pode ser negativo.")
            except Exception:
                row["erros"].append("Custo unitário inválido.")
                row["custo_unitario_dec"] = Decimal("0")
            try:
                row["estoque_minimo_int"] = max(0, int(str(row["estoque_minimo"] or "0")))
                row["estoque_inicial_int"] = max(0, int(str(row["estoque_inicial"] or "0")))
            except Exception:
                row["erros"].append("Estoque mínimo/inicial inválido.")
                row["estoque_minimo_int"] = 0
                row["estoque_inicial_int"] = 0
            if row["nome"] and Produto.objects.filter(nome__iexact=row["nome"]).exists():
                row["erros"].append("Já existe produto com este nome no sistema.")
            if row["ean"] and Produto.objects.filter(ean=row["ean"]).exists():
                row["erros"].append("Já existe produto com este EAN no sistema.")

            normalizadas.append(row)
            if row["erros"]:
                erros.append((idx, list(row["erros"])))

        preview = normalizadas[:200]
        if acao == "importar" and not erros:
            with transaction.atomic():
                for row in normalizadas:
                    marca = None
                    fornecedor = None
                    if row["marca_nome"]:
                        marca = MarcaGarantia.objects.filter(nome__iexact=row["marca_nome"], ativo=True).first()
                    if row["fornecedor_nome"]:
                        fornecedor = FornecedorGarantia.objects.filter(nome__iexact=row["fornecedor_nome"], ativo=True).first()
                    produto = Produto.objects.create(
                        nome=row["nome"],
                        sku=row["sku"] or None,
                        ean=row["ean"] or None,
                        tipo_item=row["tipo_item"],
                        categoria=row["categoria"],
                        marca=marca,
                        fornecedor_config=fornecedor,
                        custo_unitario=row["custo_unitario_dec"],
                        preco_final=row["preco_final_dec"],
                        estoque_minimo=row["estoque_minimo_int"],
                        quantidade=0,
                        ativo=True,
                    )
                    _aplicar_estoque_inicial(
                        produto,
                        estoque_inicial=row["estoque_inicial_int"],
                        usuario=request.user,
                        observacao="Entrada inicial por importação.",
                    )
                    _normalizar_saldos_produto(produto)
                    _recalcular_total_produto(produto)
                    _registrar_historico_produto(
                        produto,
                        usuario=request.user,
                        acao="IMPORTACAO",
                        dados_antes={},
                        observacao="Produto criado via importação de arquivo.",
                    )
                    importados += 1
            messages.success(request, f"Importação concluída: {importados} produto(s).")
            return redirect("estoque:lista_produtos")

        if erros:
            messages.warning(request, f"Foram encontradas {len(erros)} linha(s) com erro. Corrija antes de importar.")
        else:
            messages.info(request, "Validação concluída sem erros. Clique em importar para confirmar.")

    return render(
        request,
        "estoque/importar_produtos.html",
        {
            "preview": preview,
            "erros": erros[:50],
            "menu_app": "estoque",
            "menu_sub": "importar_produtos",
        },
    )


@role_required(STOCK_MANAGE_ROLES)
def tabelas_preco(request):
    if request.method == "POST":
        acao = (request.POST.get("acao") or "").strip()
        if acao == "excluir":
            tabela_id = request.POST.get("tabela_id")
            if tabela_id and tabela_id.isdigit():
                tabela = TabelaPreco.objects.filter(id=int(tabela_id)).first()
                if tabela:
                    tabela.delete()
                    messages.success(request, "Tabela de preco excluida.")
            return redirect("estoque:tabelas_preco")
        form = TabelaPrecoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Tabela de preco salva.")
            return redirect("estoque:tabelas_preco")
    else:
        form = TabelaPrecoForm()

    return render(
        request,
        "estoque/tabelas_preco.html",
        {
            "form": form,
            "tabelas": TabelaPreco.objects.order_by("nome"),
            "menu_app": "estoque",
            "menu_sub": "tabelas_preco",
        },
    )


@role_required(STOCK_MANAGE_ROLES)
def estrutura_produto(request, produto_id):
    produto = get_object_or_404(Produto, id=produto_id)

    if request.method == "POST":
        acao = (request.POST.get("acao") or "").strip()
        if acao == "adicionar_preco_tabela":
            preco_form = ProdutoPrecoTabelaForm(request.POST)
            if preco_form.is_valid():
                item, _ = ProdutoPrecoTabela.objects.update_or_create(
                    produto=produto,
                    tabela=preco_form.cleaned_data["tabela"],
                    defaults={"preco": preco_form.cleaned_data["preco"]},
                )
                messages.success(request, f"Preco da tabela '{item.tabela.nome}' atualizado.")
                return redirect("estoque:estrutura_produto", produto_id=produto.id)
        elif acao == "adicionar_equivalente":
            equivalente_form = ProdutoEquivalenteForm(request.POST, produto=produto)
            if equivalente_form.is_valid():
                eq = equivalente_form.save(commit=False)
                eq.produto = produto
                eq.save()
                messages.success(request, "Produto equivalente adicionado.")
                return redirect("estoque:estrutura_produto", produto_id=produto.id)
        elif acao == "adicionar_kit_item":
            kit_form = ProdutoKitItemForm(request.POST, produto=produto)
            if kit_form.is_valid():
                kit_item = kit_form.save(commit=False)
                kit_item.produto_kit = produto
                kit_item.save()
                messages.success(request, "Componente de kit adicionado.")
                return redirect("estoque:estrutura_produto", produto_id=produto.id)
        elif acao == "excluir_preco_tabela":
            item_id = request.POST.get("item_id")
            if item_id and item_id.isdigit():
                ProdutoPrecoTabela.objects.filter(id=int(item_id), produto=produto).delete()
                messages.success(request, "Preco de tabela removido.")
                return redirect("estoque:estrutura_produto", produto_id=produto.id)
        elif acao == "excluir_equivalente":
            item_id = request.POST.get("item_id")
            if item_id and item_id.isdigit():
                ProdutoEquivalente.objects.filter(id=int(item_id), produto=produto).delete()
                messages.success(request, "Equivalente removido.")
                return redirect("estoque:estrutura_produto", produto_id=produto.id)
        elif acao == "excluir_kit_item":
            item_id = request.POST.get("item_id")
            if item_id and item_id.isdigit():
                ProdutoKitItem.objects.filter(id=int(item_id), produto_kit=produto).delete()
                messages.success(request, "Componente do kit removido.")
                return redirect("estoque:estrutura_produto", produto_id=produto.id)

    preco_form = ProdutoPrecoTabelaForm()
    equivalente_form = ProdutoEquivalenteForm(produto=produto)
    kit_form = ProdutoKitItemForm(produto=produto)

    return render(
        request,
        "estoque/estrutura_produto.html",
        {
            "produto": produto,
            "preco_form": preco_form,
            "equivalente_form": equivalente_form,
            "kit_form": kit_form,
            "precos_tabela": ProdutoPrecoTabela.objects.select_related("tabela").filter(produto=produto),
            "equivalentes": ProdutoEquivalente.objects.select_related("equivalente").filter(produto=produto),
            "kit_componentes": ProdutoKitItem.objects.select_related("componente").filter(produto_kit=produto),
            "historicos": produto.historicos.select_related("usuario").all()[:20],
            "menu_app": "estoque",
            "menu_sub": "lista_produtos",
        },
    )


@role_required(STOCK_MANAGE_ROLES)
def excluir_produto(request, produto_id):
    produto = get_object_or_404(Produto, id=produto_id)
    if request.method == "POST":
        produto.delete()
        messages.success(request, "Produto excluído com sucesso!")
        return redirect("estoque:lista_produtos")
    return render(request, "estoque/confirm_delete.html", {"produto": produto, "menu_app": "estoque", "menu_sub": "lista_produtos"})


@role_required(STOCK_VIEW_ROLES)
def buscar_produto(request):
    q = (request.GET.get("q") or "").strip()
    tipo = (request.GET.get("tipo") or "").strip().lower()
    if len(q) < 2:
        return JsonResponse([], safe=False)

    produtos = Produto.objects.filter(ativo=True, permite_os=True)
    if tipo == "servico":
        produtos = produtos.filter(tipo_item="servico")
    elif tipo in {"peca", "nao_servico"}:
        produtos = produtos.exclude(tipo_item="servico")
    elif tipo in {"produto", "consumivel"}:
        produtos = produtos.filter(tipo_item=tipo)

    produtos = produtos.filter(
        Q(nome__icontains=q) | Q(ean__icontains=q) | Q(sku__icontains=q) | Q(modelos_compativeis__icontains=q)
    ).order_by("nome")
    data = list(
        produtos.values(
            "id",
            "ean",
            "sku",
            "nome",
            "descricao",
            "preco",
            "tipo_item",
            "modelos_compativeis",
            "garantia_peca_dias",
        )[:50]
    )
    return JsonResponse(data, safe=False)


@role_required(STOCK_MANAGE_ROLES)
def api_gerar_ean(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "erro": "Método inválido."}, status=405)
    produto_tmp = Produto()
    codigo = produto_tmp._gerar_codigo_ean()
    return JsonResponse({"ok": True, "ean": codigo})


@role_required(ORDER_ROLES)
def api_sugerir_pecas_os(request):
    q = (request.GET.get("q") or "").strip()
    modelo = (request.GET.get("modelo") or "").strip()
    servico = (request.GET.get("servico") or "").strip()
    defeito = (request.GET.get("defeito") or "").strip()
    tipo_equipamento = (request.GET.get("tipo_equipamento") or "").strip()
    if not modelo and not servico and not q:
        return JsonResponse({"ok": True, "resultados": []})

    base_qs = Produto.objects.filter(ativo=True, is_servico=False, permite_os=True).prefetch_related("servicos_compativeis")
    produtos = base_qs
    if q:
        produtos = produtos.filter(
            Q(nome__icontains=q) | Q(ean__icontains=q) | Q(sku__icontains=q) | Q(modelos_compativeis__icontains=q)
        )
    if modelo:
        produtos = produtos.filter(modelos_compativeis__icontains=modelo)
    if servico:
        produtos = produtos.filter(Q(servicos_compativeis__nome__icontains=servico) | Q(nome__icontains=servico)).distinct()

    candidatos = list(produtos.select_related("categoria_config").order_by("nome")[:120])
    if not candidatos:
        fallback = base_qs
        if q:
            fallback = fallback.filter(Q(nome__icontains=q) | Q(ean__icontains=q) | Q(sku__icontains=q))
        candidatos = list(fallback.select_related("categoria_config").order_by("nome")[:60])

    historico_por_nome = {}
    try:
        from ordens.models import ServicoPeca

        historico = ServicoPeca.objects.filter(tipo="peca")
        if modelo:
            historico = historico.filter(ordem__modelo_equipamento__icontains=modelo)
        if tipo_equipamento:
            historico = historico.filter(ordem__tipo_equipamento__icontains=tipo_equipamento)
        if defeito:
            historico = historico.filter(ordem__defeito__icontains=defeito)
        if servico:
            historico = historico.filter(
                Q(nome__icontains=servico) | Q(descricao__icontains=servico) | Q(ordem__relatorio_tecnico__icontains=servico)
            )

        for row in historico.values("nome").annotate(total=Count("id")):
            chave = _normalizar_texto(row.get("nome"))
            if not chave:
                continue
            atual = int(historico_por_nome.get(chave, 0))
            historico_por_nome[chave] = max(atual, int(row.get("total") or 0))
    except Exception:
        historico_por_nome = {}

    modelo_norm = _normalizar_texto(modelo)
    servico_norm = _normalizar_texto(servico)
    q_norm = _normalizar_texto(q)
    ranked = []
    for p in candidatos:
        nome_norm = _normalizar_texto(p.nome)
        modelos_norm = _normalizar_texto(p.modelos_compativeis)
        historico = int(historico_por_nome.get(nome_norm, 0))
        if historico == 0:
            for chave, total in historico_por_nome.items():
                if nome_norm and chave and (nome_norm in chave or chave in nome_norm):
                    historico = max(historico, int(total or 0))

        score = 0
        motivos = []
        if modelo_norm and modelos_norm and modelo_norm in modelos_norm:
            score += 55
            motivos.append("Modelo compatível")
        if servico_norm and any(servico_norm in _normalizar_texto(s.nome) for s in p.servicos_compativeis.all()):
            score += 24
            motivos.append("Serviço compatível")
        if q_norm and (q_norm in nome_norm or q_norm in _normalizar_texto(p.ean) or q_norm in _normalizar_texto(p.sku)):
            score += 14
            motivos.append("Termo da busca")
        if historico > 0:
            bonus_historico = min(45, historico * 9)
            score += bonus_historico
            motivos.append(f"Historico ({historico}x)")
        if int(p.quantidade or 0) > 0:
            score += 6
            motivos.append("Com estoque")
        else:
            score -= 8
        if p.garantia_peca_dias:
            score += 2

        ranked.append(
            {
                "id": p.id,
                "nome": p.nome,
                "ean": p.ean or "",
                "sku": p.sku or "",
                "preco": float(p.preco_final),
                "garantia_peca_dias": p.garantia_peca_dias or 0,
                "modelos_compativeis": p.modelos_compativeis or "",
                "quantidade": int(p.quantidade or 0),
                "score": int(score),
                "frequencia_historica": int(historico),
                "motivos": motivos[:3],
            }
        )

    ranked.sort(key=lambda x: (x["score"], x["frequencia_historica"], x["quantidade"], x["nome"]), reverse=True)
    data = ranked[:30]
    return JsonResponse({"ok": True, "resultados": data})


@role_required(STOCK_MANAGE_ROLES)
def registrar_movimentacao(request):
    if request.method == "POST":
        form = MovimentacaoEstoqueForm(request.POST)
        if form.is_valid():
            mov = form.save(commit=False)
            mov.usuario = request.user
            produto = mov.produto
            _normalizar_saldos_produto(produto)
            config = _config_sistema()

            try:
                with transaction.atomic():
                    if mov.tipo == "transferencia":
                        if int(mov.quantidade or 0) <= 0:
                            messages.error(request, "Transferência exige quantidade positiva.")
                            return redirect("estoque:registrar_movimentacao")
                        origem_saldo, _ = SaldoEstoquePonto.objects.get_or_create(produto=produto, ponto_operacional=mov.origem)
                        destino_saldo, _ = SaldoEstoquePonto.objects.get_or_create(produto=produto, ponto_operacional=mov.destino)
                        if origem_saldo.quantidade < mov.quantidade:
                            messages.error(request, "Saldo insuficiente na origem.")
                            return redirect("estoque:registrar_movimentacao")
                        origem_saldo.quantidade -= mov.quantidade
                        destino_saldo.quantidade += mov.quantidade
                        origem_saldo.save(update_fields=["quantidade"])
                        destino_saldo.save(update_fields=["quantidade"])
                    elif mov.tipo == "entrada" and mov.destino:
                        ajustar_saldo(produto, mov.destino, abs(int(mov.quantidade)))
                        mov.quantidade = abs(int(mov.quantidade))
                        custo_entrada = mov.valor_unitario_custo if mov.valor_unitario_custo is not None else produto.custo_unitario
                        qtd_entrada = abs(int(mov.quantidade))
                        qtd_anterior = max(int(produto.quantidade or 0) - qtd_entrada, 0)
                        custo_anterior = Decimal(str(produto.custo_medio or produto.custo_unitario or 0))
                        custo_entrada_dec = Decimal(str(custo_entrada or 0))
                        if (qtd_anterior + qtd_entrada) > 0:
                            custo_medio = ((custo_anterior * qtd_anterior) + (custo_entrada_dec * qtd_entrada)) / Decimal(qtd_anterior + qtd_entrada)
                            produto.custo_medio = custo_medio
                            produto.custo_unitario = custo_medio
                            produto.save(update_fields=["custo_medio", "custo_unitario"])
                    elif mov.tipo in {"ajuste", "avaria", "inventario"} and mov.origem:
                        if not (mov.observacao or "").strip():
                            messages.error(request, "Informe observação para ajuste/avaria/inventário.")
                            return redirect("estoque:registrar_movimentacao")
                        ajustar_saldo(produto, mov.origem, mov.quantidade)
                    elif mov.tipo in {"venda", "consumo_os"} and mov.origem:
                        ajustar_saldo(
                            produto,
                            mov.origem,
                            -abs(int(mov.quantidade)),
                            allow_negative=bool(config.estoque_permitir_negativo),
                        )
                        mov.quantidade = -abs(int(mov.quantidade))
                    elif mov.tipo in {"devolucao_reserva"} and mov.destino:
                        ajustar_saldo(produto, mov.destino, abs(int(mov.quantidade)))
                        mov.quantidade = abs(int(mov.quantidade))
                    mov.save()
                    _recalcular_total_produto(produto)
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect("estoque:registrar_movimentacao")

            messages.success(request, "Movimentação registrada com sucesso.")
            return redirect("estoque:movimentacoes")
    else:
        form = MovimentacaoEstoqueForm()
    return render(request, "estoque/movimentacao_form.html", {"form": form, "menu_app": "estoque", "menu_sub": "movimentacoes"})


@role_required(STOCK_VIEW_ROLES)
def listar_movimentacoes(request):
    movimentacoes = MovimentacaoEstoque.objects.select_related("produto", "origem", "destino", "usuario")
    tipo = (request.GET.get("tipo") or "").strip()
    ponto = (request.GET.get("ponto") or "").strip()
    page_number = request.GET.get("page")
    if tipo:
        movimentacoes = movimentacoes.filter(tipo=tipo)
    if ponto:
        movimentacoes = movimentacoes.filter(Q(origem_id=ponto) | Q(destino_id=ponto))
    movimentacoes = movimentacoes.order_by("-criado_em", "-id")
    movimentacoes_page = Paginator(movimentacoes, 50).get_page(page_number)
    return render(
        request,
        "estoque/movimentacoes_list.html",
        {
            "movimentacoes": movimentacoes_page,
            "movimentacoes_page": movimentacoes_page,
            "tipos_mov": MovimentacaoEstoque.TIPO_CHOICES,
            "pontos": PontoOperacional.objects.filter(ativo=True),
            "tipo_filtro": tipo,
            "ponto_filtro": ponto,
            "menu_app": "estoque",
            "menu_sub": "movimentacoes",
        },
    )


@role_required(STOCK_MANAGE_ROLES)
def pontos_operacionais(request):
    if request.method == "POST":
        form = PontoOperacionalForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Ponto operacional salvo.")
            return redirect("estoque:pontos_operacionais")
    else:
        form = PontoOperacionalForm()

    return render(
        request,
        "estoque/pontos_operacionais.html",
        {
            "form": form,
            "pontos": PontoOperacional.objects.all(),
            "menu_app": "estoque",
            "menu_sub": "pontos_operacionais",
        },
    )


@role_required(STOCK_MANAGE_ROLES)
def ubicacoes_estoque(request):
    if request.method == "POST":
        form = UbicacaoEstoqueForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Ubicação salva.")
            return redirect("estoque:ubicacoes_estoque")
    else:
        form = UbicacaoEstoqueForm()

    return render(
        request,
        "estoque/ubicacoes_estoque.html",
        {
            "form": form,
            "ubicacoes": UbicacaoEstoque.objects.select_related("ponto_operacional").all(),
            "menu_app": "estoque",
            "menu_sub": "ubicacoes_estoque",
        },
    )


@role_required(STOCK_MANAGE_ROLES)
def transferir_estoque(request):
    q = (request.GET.get("q") or "").strip()
    produtos = Produto.objects.filter(ativo=True, is_servico=False)
    if q:
        produtos = produtos.filter(Q(nome__icontains=q) | Q(ean__icontains=q) | Q(sku__icontains=q))
    produtos = produtos.order_by("nome")[:50]
    pontos = PontoOperacional.objects.filter(ativo=True).order_by("codigo")
    ubicacoes = UbicacaoEstoque.objects.select_related("ponto_operacional").filter(ativo=True).order_by(
        "ponto_operacional__codigo", "codigo"
    )

    if request.method == "POST":
        produto = get_object_or_404(Produto, id=request.POST.get("produto_id"), ativo=True)
        origem = get_object_or_404(PontoOperacional, id=request.POST.get("origem_id"), ativo=True)
        destino = get_object_or_404(PontoOperacional, id=request.POST.get("destino_id"), ativo=True)
        destino_ubicacao_id = request.POST.get("destino_ubicacao_id")
        destino_ubicacao_txt = (request.POST.get("destino_ubicacao") or "").strip()
        try:
            quantidade = int(request.POST.get("quantidade") or "0")
        except ValueError:
            quantidade = 0

        if quantidade <= 0:
            messages.error(request, "Quantidade inválida.")
            return redirect(f"{reverse('estoque:transferir_estoque')}?q={q}")
        if origem == destino:
            messages.error(request, "Origem e destino devem ser diferentes.")
            return redirect(f"{reverse('estoque:transferir_estoque')}?q={q}")
        destino_codigo = (destino.codigo or "").upper()
        if destino_codigo == "PO2" and not destino_ubicacao_id and not destino_ubicacao_txt:
            messages.error(request, "Selecione ou informe a ubicação de destino no PO2.")
            return redirect(f"{reverse('estoque:transferir_estoque')}?q={q}")

        destino_ubicacao = destino_ubicacao_txt
        if destino_ubicacao_id:
            ub = UbicacaoEstoque.objects.filter(id=destino_ubicacao_id, ativo=True).select_related("ponto_operacional").first()
            if ub:
                if ub.ponto_operacional_id != destino.id:
                    messages.error(request, "A ubicação selecionada não pertence ao ponto de destino.")
                    return redirect(f"{reverse('estoque:transferir_estoque')}?q={q}")
                destino_ubicacao = ub.codigo if not ub.descricao else f"{ub.codigo} - {ub.descricao}"

        with transaction.atomic():
            SaldoEstoquePonto.objects.get_or_create(produto=produto, ponto_operacional=origem)
            SaldoEstoquePonto.objects.get_or_create(produto=produto, ponto_operacional=destino)
            disponivel = saldo_disponivel(produto, origem)
            if disponivel < quantidade:
                messages.error(request, f"Saldo insuficiente na origem. Disponível: {disponivel}.")
                return redirect(f"{reverse('estoque:transferir_estoque')}?q={q}")
            try:
                ajustar_saldo(produto, origem, -quantidade)
                ajustar_saldo(produto, destino, quantidade)
                MovimentacaoEstoque.objects.create(
                    produto=produto,
                    tipo="transferencia",
                    quantidade=quantidade,
                    origem=origem,
                    destino=destino,
                    destino_ubicacao=destino_ubicacao,
                    observacao=f"Transpasse por busca de artigo. {destino_ubicacao}".strip(),
                    usuario=request.user,
                )
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect(f"{reverse('estoque:transferir_estoque')}?q={q}")
        logger.info(
            "transferencia_estoque",
            extra={"produto_id": produto.id, "origem_id": origem.id, "destino_id": destino.id, "quantidade": quantidade, "usuario_id": request.user.id},
        )
        messages.success(request, "Transferência registrada com sucesso.")
        return redirect("estoque:movimentacoes")

    return render(
        request,
        "estoque/transferir_estoque.html",
        {
            "produtos": produtos,
            "pontos": pontos,
            "ubicacoes": ubicacoes,
            "q": q,
            "menu_app": "estoque",
            "menu_sub": "transferir_estoque",
        },
    )


@role_required(STOCK_MANAGE_ROLES)
def reposicao_estoque(request):
    po2 = PontoOperacional.objects.filter(codigo__iexact="PO2", ativo=True).first()
    po3 = PontoOperacional.objects.filter(codigo__iexact="PO3", ativo=True).first()
    if not po2 or not po3:
        messages.error(request, "Configure os pontos PO2 (Armazém) e PO3 (Loja) para usar reposição inteligente.")
        return redirect("estoque:pontos_operacionais")

    if request.method == "POST":
        produto = get_object_or_404(Produto, id=request.POST.get("produto_id"), ativo=True, is_servico=False)
        try:
            quantidade = int(request.POST.get("quantidade") or "0")
        except ValueError:
            quantidade = 0
        if quantidade <= 0:
            messages.error(request, "Quantidade inválida para reposição.")
            return redirect("estoque:reposicao_estoque")

        with transaction.atomic():
            saldo_origem = SaldoEstoquePonto.objects.get_or_create(produto=produto, ponto_operacional=po2)[0]
            SaldoEstoquePonto.objects.get_or_create(produto=produto, ponto_operacional=po3)
            if saldo_origem.quantidade < quantidade:
                messages.error(
                    request,
                    f"Saldo insuficiente no PO2 para {produto.nome}. Disponivel: {saldo_origem.quantidade}.",
                )
                return redirect("estoque:reposicao_estoque")
            try:
                ajustar_saldo(produto, po2, -quantidade)
                ajustar_saldo(produto, po3, quantidade)
                MovimentacaoEstoque.objects.create(
                    produto=produto,
                    tipo="transferencia",
                    quantidade=quantidade,
                    origem=po2,
                    destino=po3,
                    observacao="Reposicao inteligente PO2 -> PO3",
                    usuario=request.user,
                )
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect("estoque:reposicao_estoque")
        messages.success(request, f"Reposição realizada: {quantidade} un de {produto.nome}.")
        return redirect("estoque:reposicao_estoque")

    produtos = Produto.objects.filter(ativo=True, is_servico=False).order_by("nome")
    linhas = []
    for p in produtos:
        saldo_po2 = SaldoEstoquePonto.objects.filter(produto=p, ponto_operacional=po2).values_list("quantidade", flat=True).first() or 0
        saldo_po3 = SaldoEstoquePonto.objects.filter(produto=p, ponto_operacional=po3).values_list("quantidade", flat=True).first() or 0
        minimo = int(p.estoque_minimo or 0)
        sugestao = max(minimo - int(saldo_po3), 0)
        if sugestao <= 0:
            continue
        pode_repor = max(min(sugestao, int(saldo_po2)), 0)
        faltante_compra = max(sugestao - int(saldo_po2), 0)
        linhas.append(
            {
                "produto": p,
                "saldo_po2": int(saldo_po2),
                "saldo_po3": int(saldo_po3),
                "minimo": minimo,
                "sugestao": sugestao,
                "pode_repor": pode_repor,
                "faltante_compra": faltante_compra,
            }
        )

    return render(
        request,
        "estoque/reposicao_estoque.html",
        {
            "linhas": linhas,
            "po2": po2,
            "po3": po3,
            "menu_app": "estoque",
            "menu_sub": "reposicao_estoque",
        },
    )


@role_required(STOCK_VIEW_ROLES)
def indicadores_estoque(request):
    hoje = timezone.localdate()
    corte_30 = timezone.now() - timedelta(days=30)
    corte_60 = timezone.now() - timedelta(days=60)
    rateio_config = ConfiguracaoRateioCustoFixo.get_solo()
    config_form = ConfiguracaoRateioCustoFixoForm(instance=rateio_config)
    snapshot_form = GerarSnapshotRateioForm(initial={"competencia": hoje.replace(day=1)})

    if request.method == "POST" and has_role(request.user, STOCK_MANAGE_ROLES):
        acao_rateio = request.POST.get("acao_rateio")
        if acao_rateio == "salvar_configuracao":
            config_form = ConfiguracaoRateioCustoFixoForm(request.POST, instance=rateio_config)
            if config_form.is_valid():
                config_form.save()
                messages.success(request, "Regra de rateio atualizada.")
                return redirect("estoque:indicadores_estoque")
            messages.error(request, "Revise a configuração da regra de rateio.")
        elif acao_rateio == "gerar_snapshot":
            snapshot_form = GerarSnapshotRateioForm(request.POST)
            if snapshot_form.is_valid():
                snapshot, criado = RateioCustoFixoCompetencia.gerar_snapshot(
                    competencia=snapshot_form.cleaned_data["competencia"],
                    usuario=request.user,
                    observacao=snapshot_form.cleaned_data.get("observacao", ""),
                )
                if criado:
                    messages.success(request, f"Snapshot do rateio de {snapshot.competencia:%m/%Y} gerado com sucesso.")
                else:
                    messages.warning(request, f"Ja existe snapshot fechado para {snapshot.competencia:%m/%Y}.")
                return redirect("estoque:indicadores_estoque")
            messages.error(request, "Informe uma competência válida para gerar o snapshot.")

    produtos = list(Produto.objects.filter(ativo=True, is_servico=False).order_by("nome"))

    total_itens = len(produtos)
    ruptura = 0
    abaixo_minimo = 0
    parados_60 = 0
    valor_estoque = Decimal("0.00")

    for p in produtos:
        qtd = int(p.quantidade or 0)
        minimo = int(p.estoque_minimo or 0)
        if qtd <= 0:
            ruptura += 1
        if qtd <= minimo:
            abaixo_minimo += 1
        if qtd > 0:
            valor_estoque += Decimal(str(p.preco_final or 0)) * Decimal(qtd)
        ultima_mov = p.movimentacoes.order_by("-criado_em").values_list("criado_em", flat=True).first()
        if not ultima_mov or ultima_mov < corte_60:
            parados_60 += 1

    negativos_ponto = (
        SaldoEstoquePonto.objects.select_related("produto", "ponto_operacional")
        .filter(quantidade__lt=0)
        .order_by("ponto_operacional__codigo", "produto__nome")
    )
    top_mov = (
        MovimentacaoEstoque.objects.filter(tipo__in=["venda", "consumo_os"], criado_em__gte=corte_30)
        .values("produto__nome", "produto_id")
        .annotate(total=Sum("quantidade"))
        .order_by("total")[:10]
    )
    top_saidas = [
        {
            "produto_id": r["produto_id"],
            "produto": r["produto__nome"],
            "unidades": abs(int(r["total"] or 0)),
        }
        for r in top_mov
        if int(r["total"] or 0) < 0
    ]
    rateio_resumo = _resumo_rateio_atual()

    return render(
        request,
        "estoque/indicadores_estoque.html",
        {
            "kpis": {
                "total_itens": total_itens,
                "ruptura": ruptura,
                "abaixo_minimo": abaixo_minimo,
                "parados_60": parados_60,
                "valor_estoque": valor_estoque,
            },
            "top_saidas": top_saidas,
            "negativos_ponto": negativos_ponto[:100],
            "hoje": hoje,
            "rateio_resumo": rateio_resumo,
            "rateio_config_form": config_form,
            "rateio_snapshot_form": snapshot_form,
            "pode_gerenciar_rateio": has_role(request.user, STOCK_MANAGE_ROLES),
            "menu_app": "estoque",
            "menu_sub": "indicadores_estoque",
        },
    )


@role_required(STOCK_VIEW_ROLES)
def detalhe_rateio_competencia(request, snapshot_id):
    snapshot = get_object_or_404(
        RateioCustoFixoCompetencia.objects.select_related("gerado_por"),
        pk=snapshot_id,
    )
    itens = snapshot.itens.select_related("produto").all()
    totais = {
        "faturamento_realizado": sum((item.faturamento_realizado for item in itens), Decimal("0.00")),
        "margem_realizada": sum((item.margem_realizada for item in itens), Decimal("0.00")),
        "quantidade_realizada": sum((item.quantidade_realizada for item in itens), 0),
    }
    return render(
        request,
        "estoque/rateio_competencia_detalhe.html",
        {
            "snapshot": snapshot,
            "itens": itens,
            "totais": totais,
            "menu_app": "estoque",
            "menu_sub": "indicadores_estoque",
        },
    )


@role_required(STOCK_VIEW_ROLES)
def exportar_rateio_competencia(request, snapshot_id):
    snapshot = get_object_or_404(
        RateioCustoFixoCompetencia.objects.select_related("gerado_por"),
        pk=snapshot_id,
    )
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="rateio_{snapshot.competencia:%Y_%m}.csv"'
    response.write("\ufeff")

    writer = csv.writer(response, delimiter=";")
    writer.writerow(
        [
            "competencia",
            "criterio",
            "produto",
            "previsao_venda_mensal",
            "quantidade_realizada",
            "base_rateio",
            "participacao_percentual",
            "custo_rateio_unitario",
            "custo_rateio_total",
            "preco_referencia",
            "lucro_unitario_referencia",
            "faturamento_realizado",
            "margem_realizada",
        ]
    )
    for item in snapshot.itens.select_related("produto").all():
        writer.writerow(
            [
                snapshot.competencia.strftime("%Y-%m-%d"),
                snapshot.get_criterio_rateio_display(),
                item.produto_nome,
                item.previsao_venda_mensal,
                item.quantidade_realizada,
                f"{item.base_rateio:.2f}",
                f"{item.participacao_percentual:.2f}",
                f"{item.custo_rateio_unitario:.2f}",
                f"{item.custo_rateio_total:.2f}",
                f"{item.preco_referencia:.2f}",
                f"{item.lucro_unitario_referencia:.2f}",
                f"{item.faturamento_realizado:.2f}",
                f"{item.margem_realizada:.2f}",
            ]
        )
    return response


@role_required(STOCK_VIEW_ROLES)
def exportar_rateio_competencia_excel(request, snapshot_id):
    snapshot = get_object_or_404(
        RateioCustoFixoCompetencia.objects.select_related("gerado_por"),
        pk=snapshot_id,
    )
    response = HttpResponse(content_type="application/vnd.ms-excel; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="rateio_{snapshot.competencia:%Y_%m}.xls"'
    response.write("\ufeff")

    linhas = [
        "<html><head><meta charset='utf-8'></head><body>",
        f"<h3>Snapshot de Rateio {snapshot.competencia:%m/%Y}</h3>",
        "<table border='1'>",
        "<tr>"
        "<th>Competência</th>"
        "<th>Critério</th>"
        "<th>Produto</th>"
        "<th>Previsto</th>"
        "<th>Realizado</th>"
        "<th>Base rateio</th>"
        "<th>Participação %</th>"
        "<th>Rateio unitário</th>"
        "<th>Rateio total</th>"
        "<th>Preço referência</th>"
        "<th>Lucro referência</th>"
        "<th>Faturamento realizado</th>"
        "<th>Margem realizada</th>"
        "</tr>",
    ]
    for item in snapshot.itens.select_related("produto").all():
        linhas.append(
            "<tr>"
            f"<td>{snapshot.competencia:%Y-%m-%d}</td>"
            f"<td>{snapshot.get_criterio_rateio_display()}</td>"
            f"<td>{item.produto_nome}</td>"
            f"<td>{item.previsao_venda_mensal}</td>"
            f"<td>{item.quantidade_realizada}</td>"
            f"<td>{item.base_rateio:.2f}</td>"
            f"<td>{item.participacao_percentual:.2f}</td>"
            f"<td>{item.custo_rateio_unitario:.2f}</td>"
            f"<td>{item.custo_rateio_total:.2f}</td>"
            f"<td>{item.preco_referencia:.2f}</td>"
            f"<td>{item.lucro_unitario_referencia:.2f}</td>"
            f"<td>{item.faturamento_realizado:.2f}</td>"
            f"<td>{item.margem_realizada:.2f}</td>"
            "</tr>"
        )
    linhas.append("</table></body></html>")
    response.write("".join(linhas))
    return response


@role_required(STOCK_MANAGE_ROLES)
def relatorio_divergencias_estoque(request):
    hoje = timezone.localdate()
    pre_reservas_antigas = VendaRapidaEstoque.objects.filter(status="pre_reserva", criado_em__date__lt=hoje)
    total_pre_reservas_antigas = pre_reservas_antigas.count()
    reservas_vencidas_ativas = ReservaEstoque.objects.filter(status="ativa", valido_ate__lt=hoje)
    produtos_abaixo = [
        p for p in Produto.objects.filter(ativo=True, is_servico=False).order_by("nome")
        if int(p.quantidade) <= int(p.estoque_minimo or 0)
    ]
    negativos_ponto = (
        SaldoEstoquePonto.objects.select_related("produto", "ponto_operacional")
        .filter(quantidade__lt=0)
        .order_by("ponto_operacional__codigo", "produto__nome")
    )
    po2 = PontoOperacional.objects.filter(codigo__iexact="PO2").first()
    mov_po2_sem_ubicacao = MovimentacaoEstoque.objects.none()
    if po2:
        mov_po2_sem_ubicacao = (
            MovimentacaoEstoque.objects.select_related("produto", "origem", "destino")
            .filter(tipo="transferencia", destino=po2)
            .filter(Q(destino_ubicacao__isnull=True) | Q(destino_ubicacao__exact=""))
            .order_by("-criado_em")
        )
    return render(
        request,
        "estoque/relatorio_divergencias.html",
        {
            "pre_reservas_antigas": pre_reservas_antigas[:200],
            "total_pre_reservas_antigas": total_pre_reservas_antigas,
            "dias_limpeza_pre_reserva": 1,
            "reservas_vencidas_ativas": reservas_vencidas_ativas[:200],
            "produtos_abaixo": produtos_abaixo[:200],
            "negativos_ponto": negativos_ponto[:200],
            "mov_po2_sem_ubicacao": mov_po2_sem_ubicacao[:200],
            "menu_app": "estoque",
            "menu_sub": "relatorio_divergencias",
        },
    )


@role_required(ORDER_ROLES)
def consulta_artigos(request):
    user_model = get_user_model()
    tecnicos_qs = (
        user_model.objects.filter(is_active=True)
        .exclude(numero_vendedor__isnull=True)
        .exclude(numero_vendedor="")
        .order_by("username")
    )
    if not tecnicos_qs.exists():
        tecnicos_qs = (
            user_model.objects.filter(is_active=True, tipo_usuario="tecnico")
            .order_by("username")
        )
    tecnicos = tecnicos_qs.values("id", "username", "numero_vendedor")
    return render(
        request,
        "estoque/consulta_artigos.html",
        {
            "menu_app": "estoque",
            "menu_sub": "consulta_artigos",
            "numero_vendedor_padrao": (getattr(request.user, "numero_vendedor", "") or ""),
            "tecnicos_disponiveis": list(tecnicos),
            "pode_venda_mostrador": has_role(request.user, STOCK_MANAGE_ROLES),
        },
    )


@role_required(ORDER_ROLES)
def api_consulta_artigos(request):
    q = (request.GET.get("q") or "").strip()
    try:
        page = max(1, int(request.GET.get("page") or "1"))
    except ValueError:
        page = 1
    page_size = 20

    # Evita varredura geral sem termo para reduzir carga.
    if len(q) < 2:
        return JsonResponse(
            {
                "resultados": [],
                "page": 1,
                "has_next": False,
                "has_prev": False,
                "total": 0,
            }
        )

    produtos = Produto.objects.filter(ativo=True, is_servico=False, permite_os=True)
    q_low = q.lower()
    if q_low.isdigit():
        produtos = produtos.filter(
            Q(id=int(q_low))
            | Q(ean__icontains=q)
            | Q(sku__icontains=q)
            | Q(nome__icontains=q)
            | Q(modelos_compativeis__icontains=q)
        )
    else:
        produtos = produtos.filter(
            Q(nome__icontains=q) | Q(ean__icontains=q) | Q(sku__icontains=q) | Q(modelos_compativeis__icontains=q)
        )

    inicio = (page - 1) * page_size
    fim = inicio + page_size
    total = produtos.count()
    data = [
        {
            "id": p.id,
            "nome": p.nome,
            "descricao": p.descricao or "",
            "ean": p.ean or "",
            "sku": p.sku or "",
            "preco": float(p.preco_final),
            "quantidade": p.quantidade,
            "modelos_compativeis": p.modelos_compativeis or "",
        }
        for p in produtos.order_by("nome")[inicio:fim]
    ]
    return JsonResponse(
        {
            "resultados": data,
            "page": page,
            "has_next": total > fim,
            "has_prev": page > 1,
            "total": total,
        }
    )


@role_required(ORDER_ROLES)
def api_resumo_artigo(request, produto_id):
    expirar_reservas_vencidas()
    produto = get_object_or_404(Produto, id=produto_id, ativo=True)
    _normalizar_saldos_produto(produto)

    pontos = list(PontoOperacional.objects.filter(ativo=True).order_by("codigo"))
    saldos_map = {s.ponto_operacional_id: s.quantidade for s in produto.saldos_por_ponto.select_related("ponto_operacional")}

    reservas_ativas = (
        ReservaEstoque.objects.filter(produto=produto, status="ativa", valido_ate__gte=timezone.localdate())
        .values("ponto_operacional_id")
        .annotate(total=Sum("quantidade"))
    )
    reservas_map = {r["ponto_operacional_id"]: int(r["total"] or 0) for r in reservas_ativas}

    estoque_pontos = []
    for p in pontos:
        qtd = int(saldos_map.get(p.id, 0))
        reservado = int(reservas_map.get(p.id, 0))
        estoque_pontos.append(
            {
                "id": p.id,
                "codigo": p.codigo,
                "nome": p.nome,
                "quantidade": qtd,
                "reservado": reservado,
                "disponivel": (qtd - reservado),
            }
        )

    reservas_recentes = [
        {
            "codigo": r.codigo_reserva,
            "nome": r.nome_contato,
            "telefone": r.telefone_contato,
            "quantidade": r.quantidade,
            "valido_ate": r.valido_ate.strftime("%d/%m/%Y"),
            "status": r.status,
            "ponto": r.ponto_operacional.codigo if r.ponto_operacional else "-",
        }
        for r in ReservaEstoque.objects.filter(produto=produto).select_related("ponto_operacional")[:15]
    ]

    movimentacoes_recentes = [
        {
            "tipo": m.get_tipo_display(),
            "quantidade": m.quantidade,
            "origem": m.origem.codigo if m.origem else "-",
            "destino": m.destino.codigo if m.destino else "-",
            "quando": timezone.localtime(m.criado_em).strftime("%d/%m/%Y %H:%M"),
            "obs": m.observacao or "",
        }
        for m in produto.movimentacoes.select_related("origem", "destino").all()[:20]
    ]

    return JsonResponse(
        {
            "id": produto.id,
            "nome": produto.nome,
            "ean": produto.ean or "",
            "sku": produto.sku or "",
            "descricao": produto.descricao or "",
            "observacao_interna": produto.observacao_interna or "",
            "localizacao": produto.localizacao or "",
            "garantia_peca_dias": produto.garantia_peca_dias or 0,
            "modelos_compativeis": produto.modelos_compativeis or "",
            "preco": float(produto.preco_final),
            "quantidade_total": produto.quantidade,
            "estoque_minimo": produto.estoque_minimo,
            "abaixo_minimo": produto.quantidade <= int(produto.estoque_minimo or 0),
            "ponto_padrao_id": produto.ponto_operacional_id,
            "estoque_pontos": estoque_pontos,
            "reservas": reservas_recentes,
            "movimentacoes": movimentacoes_recentes,
        }
    )


@role_required(STOCK_MANAGE_ROLES)
def api_venda_rapida(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "erro": "Método inválido."}, status=405)

    produto_id = request.POST.get("produto_id")
    ponto_id = request.POST.get("ponto_id")
    funcionario_numero = (request.POST.get("funcionario_numero") or "").strip()
    cesto_codigo = (request.POST.get("cesto_codigo") or "").strip()

    try:
        quantidade = int(request.POST.get("quantidade") or "1")
    except ValueError:
        return JsonResponse({"ok": False, "erro": "Quantidade inválida."}, status=400)

    if quantidade <= 0:
        return JsonResponse({"ok": False, "erro": "Quantidade deve ser maior que zero."}, status=400)
    if not funcionario_numero.isdigit() or len(funcionario_numero) < 2:
        return JsonResponse({"ok": False, "erro": "Número de vendedor inválido. Use ao menos 2 dígitos."}, status=400)

    produto = get_object_or_404(Produto, id=produto_id, ativo=True, is_servico=False, permite_os=True)
    ponto = get_object_or_404(PontoOperacional, id=ponto_id, ativo=True)
    _normalizar_saldos_produto(produto)
    config = _config_sistema()

    codigo_ref = ponto.codigo.upper()
    if codigo_ref not in {"PO3", "PO2"}:
        return JsonResponse({"ok": False, "erro": "Venda permitida apenas para pontos PO3 (Loja) e PO2 (Armazem)."}, status=400)
    if not get_user_model().objects.filter(is_active=True, numero_vendedor=funcionario_numero).exists():
        return JsonResponse({"ok": False, "erro": "Numero de vendedor nao encontrado para usuario ativo."}, status=400)

    with transaction.atomic():
        if cesto_codigo:
            cesto_em_aberto = VendaRapidaEstoque.objects.filter(cesto_codigo=cesto_codigo, status="pre_reserva")
            if cesto_em_aberto.exclude(guia_pagamento="").exists():
                return JsonResponse(
                    {
                        "ok": False,
                        "erro": "Este cesto ja foi finalizado. Inicie um novo cesto para continuar.",
                    },
                    status=409,
                )
        SaldoEstoquePonto.objects.get_or_create(produto=produto, ponto_operacional=ponto)
        pre_reservado = (
            VendaRapidaEstoque.objects.filter(
                produto=produto,
                ponto_operacional=ponto,
                status="pre_reserva",
            ).aggregate(total=Sum("quantidade"))["total"]
            or 0
        )
        if config.estoque_pre_reserva_exige_saldo:
            saldo_atual = SaldoEstoquePonto.objects.get(produto=produto, ponto_operacional=ponto)
            disponivel = int(saldo_atual.quantidade) - int(pre_reservado)
            if disponivel < quantidade:
                return JsonResponse(
                    {
                        "ok": False,
                        "erro": (
                            f"Saldo insuficiente para pre-reserva no ponto {ponto.codigo}. "
                            f"Disponivel: {disponivel}."
                        ),
                    },
                    status=400,
                )
        valor_unitario = Decimal(str(produto.preco_final))
        valor_total = valor_unitario * quantidade

        if not cesto_codigo:
            cesto_codigo = _codigo_cesto()

        venda = VendaRapidaEstoque.objects.create(
            produto=produto,
            ponto_operacional=ponto,
            quantidade=quantidade,
            valor_unitario=valor_unitario,
            valor_total=valor_total,
            funcionario_numero=funcionario_numero,
            cesto_codigo=cesto_codigo,
            status="pre_reserva",
            usuario=request.user,
        )
        total_cesto = (
            VendaRapidaEstoque.objects.filter(cesto_codigo=cesto_codigo, status="pre_reserva")
            .aggregate(total=Sum("valor_total"))["total"]
            or Decimal("0.00")
        )
    logger.info(
        "venda_pre_reserva_criada",
        extra={"venda_id": venda.id, "produto_id": produto.id, "ponto_id": ponto.id, "quantidade": quantidade, "usuario_id": request.user.id},
    )
    return JsonResponse(
        {
            "ok": True,
            "venda_id": venda.id,
            "cesto_codigo": cesto_codigo,
            "valor_total": float(venda.valor_total),
            "total_cesto": float(total_cesto),
        }
    )


@role_required(STOCK_MANAGE_ROLES)
def api_cesto_resumo(request, cesto_codigo):
    return JsonResponse(_resumo_cesto(cesto_codigo))


@role_required(STOCK_MANAGE_ROLES)
def api_cesto_finalizar(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "erro": "Método inválido."}, status=405)

    cesto_codigo = (request.POST.get("cesto_codigo") or "").strip()
    if not cesto_codigo:
        return JsonResponse({"ok": False, "erro": "Cesto inválido."}, status=400)

    vendas_qs = VendaRapidaEstoque.objects.filter(cesto_codigo=cesto_codigo, status="pre_reserva")
    if not vendas_qs.exists():
        return JsonResponse({"ok": False, "erro": "Cesto vazio ou já finalizado."}, status=400)

    guia = vendas_qs.exclude(guia_pagamento="").values_list("guia_pagamento", flat=True).first()
    if not guia:
        guia = _codigo_guia()
    vendas_qs.exclude(guia_pagamento=guia).update(guia_pagamento=guia)

    resumo = _resumo_cesto(cesto_codigo)
    return JsonResponse(
        {
            "ok": True,
            "guia": guia,
            "total": resumo["total"],
            "itens": len(resumo["itens"]),
            "redirect_caixa": f"{reverse('caixa:registrar_pagamento')}?guia={guia}",
            "imprimir_url": reverse("estoque:guia_pagamento", args=[guia]),
        }
    )


@role_required(STOCK_MANAGE_ROLES)
def api_cesto_item_remover(request, venda_id):
    if request.method != "POST":
        return JsonResponse({"ok": False, "erro": "Método inválido."}, status=405)

    cesto_codigo = (request.POST.get("cesto_codigo") or "").strip()
    if not cesto_codigo:
        return JsonResponse({"ok": False, "erro": "Informe o código do cesto."}, status=400)
    venda = get_object_or_404(VendaRapidaEstoque, id=venda_id)

    if venda.status != "pre_reserva":
        return JsonResponse({"ok": False, "erro": "Somente itens em pré-reserva podem ser removidos."}, status=400)
    if cesto_codigo and venda.cesto_codigo != cesto_codigo:
        return JsonResponse({"ok": False, "erro": "Item não pertence ao cesto informado."}, status=400)

    venda.status = "cancelada"
    venda.concluido_em = timezone.now()
    venda.save(update_fields=["status", "concluido_em"])

    return JsonResponse(_resumo_cesto(venda.cesto_codigo))


@role_required(CAIXA_OPERATIONAL_ROLES)
def guia_pagamento(request, guia_codigo):
    vendas_qs = (
        VendaRapidaEstoque.objects.select_related("produto", "ponto_operacional", "usuario")
        .filter(guia_pagamento=guia_codigo)
        .order_by("id")
    )
    if not vendas_qs.exists():
        messages.error(request, "Guia não encontrada.")
        return redirect("estoque:consulta_artigos")
    vendas = list(vendas_qs)
    user_model = get_user_model()
    numeros = [v.funcionario_numero for v in vendas if v.funcionario_numero]
    tecnicos_map = {
        u.numero_vendedor: u.username
        for u in user_model.objects.filter(numero_vendedor__in=numeros, is_active=True)
    }
    for venda in vendas:
        venda.tecnico_nome = tecnicos_map.get(venda.funcionario_numero, "-")
    total = sum((v.valor_total for v in vendas), Decimal("0.00"))
    return render(
        request,
        "estoque/guia_pagamento.html",
        {
            "guia_codigo": guia_codigo,
            "vendas": vendas,
            "total": total,
            "menu_app": "estoque",
            "menu_sub": "consulta_artigos",
        },
    )


@role_required(ORDER_ROLES)
def api_criar_reserva(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "erro": "Método inválido."}, status=405)

    produto = get_object_or_404(Produto, id=request.POST.get("produto_id"), ativo=True, is_servico=False, permite_os=True)
    ponto = get_object_or_404(PontoOperacional, id=request.POST.get("ponto_id"), ativo=True)
    nome = (request.POST.get("nome") or "").strip()
    telefone = (request.POST.get("telefone") or "").strip()

    try:
        quantidade = int(request.POST.get("quantidade") or "1")
    except ValueError:
        return JsonResponse({"ok": False, "erro": "Quantidade inválida."}, status=400)
    if not nome:
        return JsonResponse({"ok": False, "erro": "Informe nome para reserva."}, status=400)

    valido_ate_raw = (request.POST.get("valido_ate") or "").strip()
    try:
        valido_ate = datetime.strptime(valido_ate_raw, "%Y-%m-%d").date()
    except Exception:
        return JsonResponse({"ok": False, "erro": "Data de validade inválida. Use YYYY-MM-DD."}, status=400)

    if quantidade <= 0:
        return JsonResponse({"ok": False, "erro": "Quantidade inválida."}, status=400)
    if valido_ate < timezone.localdate():
        return JsonResponse({"ok": False, "erro": "Data de validade da reserva não pode ser passada."}, status=400)

    expirar_reservas_vencidas()
    _normalizar_saldos_produto(produto)
    with transaction.atomic():
        saldo = (
            SaldoEstoquePonto.objects.select_for_update()
            .filter(produto=produto, ponto_operacional=ponto)
            .first()
        )
        if not saldo:
            saldo = SaldoEstoquePonto.objects.create(produto=produto, ponto_operacional=ponto, quantidade=0)

        reservado = (
            ReservaEstoque.objects.select_for_update()
            .filter(
                produto=produto,
                ponto_operacional=ponto,
                status="ativa",
                valido_ate__gte=timezone.localdate(),
            )
            .aggregate(total=Sum("quantidade"))["total"]
            or 0
        )
        disponivel = int(saldo.quantidade) - int(reservado)
        if disponivel < quantidade:
            return JsonResponse({"ok": False, "erro": "Sem saldo disponivel para reservar neste ponto."}, status=400)

        reserva = ReservaEstoque.objects.create(
            codigo_reserva=_codigo_reserva(),
            produto=produto,
            ponto_operacional=ponto,
            quantidade=quantidade,
            nome_contato=nome,
            telefone_contato=telefone,
            valido_ate=valido_ate,
            status="ativa",
            usuario=request.user,
        )
    logger.info(
        "reserva_criada",
        extra={"reserva_id": reserva.id, "produto_id": produto.id, "ponto_id": ponto.id, "quantidade": quantidade, "usuario_id": request.user.id},
    )

    return JsonResponse({"ok": True, "codigo_reserva": reserva.codigo_reserva})


@role_required(STOCK_MANAGE_ROLES)
def api_expirar_reservas(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "erro": "Método inválido."}, status=405)
    total = expirar_reservas_vencidas(usuario=request.user)
    logger.info("reservas_expiradas_execucao", extra={"quantidade": total, "usuario_id": request.user.id})
    return JsonResponse({"ok": True, "reservas_expiradas": total})


@role_required(STOCK_MANAGE_ROLES)
def api_converter_reserva(request, codigo_reserva):
    if request.method != "POST":
        return JsonResponse({"ok": False, "erro": "Método inválido."}, status=405)
    reserva = get_object_or_404(ReservaEstoque, codigo_reserva=codigo_reserva)
    try:
        converter_reserva(reserva, usuario=request.user, motivo="Conversao manual")
        logger.info("reserva_convertida", extra={"reserva_id": reserva.id, "usuario_id": request.user.id})
        return JsonResponse({"ok": True})
    except ValueError as exc:
        return JsonResponse({"ok": False, "erro": str(exc)}, status=400)


@role_required(STOCK_MANAGE_ROLES)
def api_cancelar_reserva(request, codigo_reserva):
    if request.method != "POST":
        return JsonResponse({"ok": False, "erro": "Método inválido."}, status=405)
    motivo = (request.POST.get("motivo") or "").strip() or "Cancelada manualmente"
    reserva = get_object_or_404(ReservaEstoque, codigo_reserva=codigo_reserva)
    try:
        cancelar_reserva(reserva, usuario=request.user, motivo=motivo)
        logger.info("reserva_cancelada", extra={"reserva_id": reserva.id, "usuario_id": request.user.id})
        return JsonResponse({"ok": True})
    except ValueError as exc:
        return JsonResponse({"ok": False, "erro": str(exc)}, status=400)


@role_required(STOCK_MANAGE_ROLES)
def api_inventario_iniciar(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "erro": "Método inválido."}, status=405)
    ponto = get_object_or_404(PontoOperacional, id=request.POST.get("ponto_id"), ativo=True)
    inventario_aberto = (
        InventarioEstoque.objects.filter(ponto_operacional=ponto, status="aberto")
        .order_by("-id")
        .first()
    )
    if inventario_aberto:
        return JsonResponse(
            {
                "ok": False,
                "erro": "Já existe inventário aberto para este ponto operacional.",
                "inventario_id": inventario_aberto.id,
            },
            status=409,
        )
    inventario = InventarioEstoque.objects.create(
        ponto_operacional=ponto,
        observacao=(request.POST.get("observacao") or "").strip(),
        usuario=request.user,
    )
    return JsonResponse({"ok": True, "inventario_id": inventario.id})


@role_required(STOCK_MANAGE_ROLES)
def api_inventario_adicionar_item(request, inventario_id):
    if request.method != "POST":
        return JsonResponse({"ok": False, "erro": "Método inválido."}, status=405)
    inventario = get_object_or_404(InventarioEstoque, id=inventario_id)
    if inventario.status != "aberto":
        return JsonResponse({"ok": False, "erro": "Inventário já finalizado."}, status=400)
    produto = get_object_or_404(Produto, id=request.POST.get("produto_id"), ativo=True, is_servico=False)
    try:
        quantidade_contada = int(request.POST.get("quantidade_contada") or "0")
    except ValueError:
        return JsonResponse({"ok": False, "erro": "Quantidade inválida."}, status=400)
    if quantidade_contada < 0:
        return JsonResponse({"ok": False, "erro": "Quantidade contada não pode ser negativa."}, status=400)

    with transaction.atomic():
        saldo = (
            SaldoEstoquePonto.objects.select_for_update()
            .filter(produto=produto, ponto_operacional=inventario.ponto_operacional)
            .first()
        )
        if not saldo:
            saldo = SaldoEstoquePonto.objects.create(produto=produto, ponto_operacional=inventario.ponto_operacional, quantidade=0)
        item, _ = ItemInventarioEstoque.objects.get_or_create(
            inventario=inventario,
            produto=produto,
            defaults={"quantidade_sistema": saldo.quantidade},
        )
        item.quantidade_sistema = saldo.quantidade
        item.quantidade_contada = quantidade_contada
        item.ajuste = quantidade_contada - saldo.quantidade
        item.observacao = (request.POST.get("observacao") or "").strip()
        item.save()
    return JsonResponse({"ok": True, "ajuste": item.ajuste})


@role_required(STOCK_MANAGE_ROLES)
def api_inventario_finalizar(request, inventario_id):
    if request.method != "POST":
        return JsonResponse({"ok": False, "erro": "Método inválido."}, status=405)
    inventario = get_object_or_404(InventarioEstoque, id=inventario_id)
    if inventario.status != "aberto":
        return JsonResponse({"ok": False, "erro": "Inventário já finalizado."}, status=400)

    itens_ajustados = 0
    unidades_ajustadas = 0
    try:
        with transaction.atomic():
            inventario = InventarioEstoque.objects.select_for_update().get(id=inventario.id)
            if inventario.status != "aberto":
                return JsonResponse({"ok": False, "erro": "Inventário já finalizado."}, status=400)

            itens = list(
                ItemInventarioEstoque.objects.select_for_update()
                .filter(inventario=inventario)
                .select_related("produto")
            )
            if not itens:
                return JsonResponse({"ok": False, "erro": "Inventário sem itens para finalizar."}, status=400)

            for item in itens:
                if item.ajuste == 0:
                    continue
                ajustar_saldo(item.produto, inventario.ponto_operacional, item.ajuste)
                MovimentacaoEstoque.objects.create(
                    produto=item.produto,
                    tipo="inventario",
                    quantidade=item.ajuste,
                    origem=inventario.ponto_operacional if item.ajuste < 0 else None,
                    destino=inventario.ponto_operacional if item.ajuste > 0 else None,
                    observacao=(
                        f"Ajuste inventario #{inventario.id} "
                        f"(sistema={item.quantidade_sistema}, contado={item.quantidade_contada}). "
                        f"{(item.observacao or '').strip()}"
                    ).strip(),
                    usuario=request.user,
                )
                itens_ajustados += 1
                unidades_ajustadas += abs(int(item.ajuste))

            inventario.status = "fechado"
            inventario.fechado_em = timezone.now()
            inventario.save(update_fields=["status", "fechado_em"])
    except ValueError as exc:
        return JsonResponse({"ok": False, "erro": str(exc)}, status=400)
    logger.info("inventario_finalizado", extra={"inventario_id": inventario.id, "usuario_id": request.user.id})
    return JsonResponse(
        {
            "ok": True,
            "resumo": {
                "itens_ajustados": itens_ajustados,
                "unidades_ajustadas": unidades_ajustadas,
            },
        }
    )


@role_required(STOCK_VIEW_ROLES)
def api_alertas_estoque(request):
    produtos = Produto.objects.filter(ativo=True, is_servico=False).order_by("nome")
    abaixo = []
    for p in produtos:
        if int(p.quantidade) <= int(p.estoque_minimo or 0):
            abaixo.append(
                {
                    "id": p.id,
                    "nome": p.nome,
                    "sku": p.sku or "",
                    "ean": p.ean or "",
                    "quantidade": int(p.quantidade),
                    "estoque_minimo": int(p.estoque_minimo or 0),
                }
            )
    return JsonResponse({"resultados": abaixo[:100]})


@role_required(STOCK_VIEW_ROLES)
def reservas_clientes(request):
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()
    page_number = request.GET.get("page")
    reservas = ReservaEstoque.objects.select_related("produto", "ponto_operacional", "ordem_servico")
    if q:
        reservas = reservas.filter(
            Q(codigo_reserva__icontains=q)
            | Q(nome_contato__icontains=q)
            | Q(telefone_contato__icontains=q)
            | Q(produto__nome__icontains=q)
        )
    if status:
        reservas = reservas.filter(status=status)
    reservas = reservas.order_by("-criado_em", "-id")
    reservas_page = Paginator(reservas, 40).get_page(page_number)
    return render(
        request,
        "estoque/reservas_clientes.html",
        {
            "reservas": reservas_page,
            "reservas_page": reservas_page,
            "q": q,
            "status_filtro": status,
            "status_choices": ReservaEstoque.STATUS_CHOICES,
            "can_manage": has_role(request.user, STOCK_MANAGE_ROLES),
            "menu_app": "estoque",
            "menu_sub": "reservas_clientes",
        },
    )


@role_required(STOCK_MANAGE_ROLES)
def associar_reserva_ordem(request, codigo_reserva):
    if request.method != "POST":
        return redirect("estoque:reservas_clientes")
    reserva = get_object_or_404(ReservaEstoque, codigo_reserva=codigo_reserva)
    ordem_id = request.POST.get("ordem_id")
    if not ordem_id:
        messages.error(request, "Informe o número da ordem (ID).")
        return redirect("estoque:reservas_clientes")
    from ordens.models import OrdemServico

    ordem = OrdemServico.objects.filter(id=ordem_id).first()
    if not ordem:
        messages.error(request, "Ordem não encontrada.")
        return redirect("estoque:reservas_clientes")
    reserva.ordem_servico = ordem
    reserva.save(update_fields=["ordem_servico"])
    messages.success(request, f"Reserva {reserva.codigo_reserva} associada a OS {ordem.numero_os}.")
    return redirect("estoque:reservas_clientes")


@role_required(STOCK_MANAGE_ROLES)
def expirar_reservas_web(request):
    if request.method == "POST":
        total = expirar_reservas_vencidas(usuario=request.user)
        messages.success(request, f"Reservas expiradas: {total}.")
    return redirect("estoque:reservas_clientes")


@role_required(STOCK_MANAGE_ROLES)
def limpar_pre_reservas_antigas_web(request):
    if request.method == "POST":
        try:
            dias = int(request.POST.get("dias") or "1")
        except ValueError:
            dias = 1
        total = limpar_pre_reservas_antigas(dias=dias)
        messages.success(request, f"Pré-reservas antigas limpas: {total}.")
    return redirect("estoque:relatorio_divergencias")


@role_required(STOCK_MANAGE_ROLES)
def converter_reserva_web(request, codigo_reserva):
    if request.method == "POST":
        reserva = get_object_or_404(ReservaEstoque, codigo_reserva=codigo_reserva)
        try:
            converter_reserva(reserva, usuario=request.user, motivo="Conversao manual")
            messages.success(request, f"Reserva {reserva.codigo_reserva} convertida.")
        except ValueError as exc:
            messages.error(request, str(exc))
    return redirect("estoque:reservas_clientes")


@role_required(STOCK_MANAGE_ROLES)
def cancelar_reserva_web(request, codigo_reserva):
    if request.method == "POST":
        reserva = get_object_or_404(ReservaEstoque, codigo_reserva=codigo_reserva)
        try:
            cancelar_reserva(reserva, usuario=request.user, motivo="Cancelamento manual")
            messages.success(request, f"Reserva {reserva.codigo_reserva} cancelada.")
        except ValueError as exc:
            messages.error(request, str(exc))
    return redirect("estoque:reservas_clientes")


def integrar_reservas_no_fechamento(ordem, usuario=None):
    return consumir_reservas_ordem(ordem, usuario=usuario)


def integrar_reservas_na_reabertura(ordem, usuario=None):
    return devolver_reservas_ordem(ordem, usuario=usuario)



